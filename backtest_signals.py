#!/usr/bin/env python3
"""
基金持仓信号回测脚本
分析公募基金季度持仓信号（加减仓、集中度）对后续个股走势的影响
输出：fund_backtest.html
"""

import json
import re
import warnings
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl

warnings.filterwarnings('ignore')

BASE   = Path("workwork/research/策略研究/基金重仓看板")
EXCEL  = BASE / "20260122 基金持仓个股视角.xlsx"
CSV_A  = BASE / "20260427 A股日收盘价.csv"
CSV_H  = BASE / "20260427 H股日收盘价.csv"
CSV_I  = BASE / "20260427 指数收盘点位.csv"
OUTPUT = BASE / "fund_backtest.html"

# ── 披露日（约季末后次月20日，若非交易日取下一交易日）──────────────────────────
DISCLOSURE_DATES = {
    '1Q24': '2024-04-20', '2Q24': '2024-07-20',
    '3Q24': '2024-10-20', '4Q24': '2025-01-20',
    '1Q25': '2025-04-20', '2Q25': '2025-07-20',
    '3Q25': '2025-10-20', '4Q25': '2026-01-20',
    '1Q26': '2026-04-20',
}

HOLD_DAYS  = [30, 60, 90]   # 持仓天数（自然日）
BENCHMARK  = '000300.SH'    # 沪深300
TOP_PCTS   = [0.10, 0.20, 0.30]  # 加减仓分组：前10%/20%/30%

# ── Excel列映射（与build_dashboard.py保持一致） ─────────────────────────────
COL = {
    'code':0,'name':1,'end_price':2,'shares':3,'shares_chg':4,
    'pos_value':5,'fund_weight':6,'float_pct':7,'chg_pct':8,'val_chg':9,
    'industry':11,'cur_price':13,'price_chg':14,'price_pct':15,
    'r60':16,'ytd':17,'float_cap':18,'fund_count':23,'fund_chg':24,
}
COL_OLD = {
    'name':0,'shares':1,'float_pct':2,'shares_chg':3,'chg_pct':4,
    'pos_value':5,'code':6,'industry':7,'fund_count':8,'fund_chg':9,
}
QUARTER_RE = re.compile(r'^([1-4]Q\d{2})\s*个股\s*$')


# ═══════════════════════════════════════════════════════════════════════════════
# 数据加载
# ═══════════════════════════════════════════════════════════════════════════════

def pv(v):
    if v is None: return None
    try:
        f = float(v); return None if f != f else round(f, 6)
    except: return None

def sort_q(q):
    m = re.match(r'(\d)Q(\d{2})', q)
    return (int(m.group(2)), int(m.group(1))) if m else (0, 0)

def _is_old(ws):
    cell = next(ws.iter_rows(min_row=1, max_row=1, max_col=1, values_only=True), [None])[0]
    return isinstance(cell, str) and '名称' in str(cell)


def load_holdings():
    print("  加载基金持仓 Excel...")
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    by_q = {}
    for sname in wb.sheetnames:
        m = QUARTER_RE.match(sname)
        if not m: continue
        q = m.group(1)
        if q in ('3Q23', '4Q23'): continue
        ws = wb[sname]; recs = []
        if _is_old(ws):
            for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
                code = row[COL_OLD['code']] if len(row) > COL_OLD['code'] else None
                if not code or not isinstance(code, str) or '.' not in str(code): continue
                g = lambda k: row[COL_OLD[k]] if len(row) > COL_OLD[k] else None
                fp = pv(g('float_pct'))
                recs.append({'code':str(code).strip(),'name':str(g('name') or '').strip(),
                    'ind':str(g('industry') or '其他').strip() or '其他',
                    'pv':pv(g('pos_value')),'fw':None,'fp':round(fp/100,6) if fp else None,
                    'fc':pv(g('fund_count')),'fcc':pv(g('fund_chg'))})
        else:
            cm = dict(COL)
            hdr = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
            if hdr:
                for i, c in enumerate(hdr):
                    if c == '持有基金数': cm['fund_count'] = i
                    elif c == '基金增减数量': cm['fund_chg'] = i
            mc = max(cm.values()) + 1
            for row in ws.iter_rows(min_row=3, max_col=mc, values_only=True):
                code = row[cm['code']] if len(row) > cm['code'] else None
                if not code or not isinstance(code, str) or '.' not in str(code): continue
                g = lambda k: row[cm[k]] if len(row) > cm[k] else None
                recs.append({'code':str(code).strip(),'name':str(g('name') or '').strip(),
                    'ind':str(g('industry') or '其他').strip() or '其他',
                    'pv':pv(g('pos_value')),'fw':pv(g('fund_weight')),
                    'fp':pv(g('float_pct')),'fc':pv(g('fund_count')),'fcc':pv(g('fund_chg'))})
        by_q[q] = recs
    wb.close()
    return by_q


def load_prices():
    print("  加载收盘价...")
    def parse(path):
        raw = pd.read_csv(path, header=None, low_memory=False)
        codes = [str(c) for c in raw.iloc[4].tolist()]
        data  = raw.iloc[5:].copy()
        data.columns = codes
        data = data.rename(columns={codes[0]: 'date'})
        data = data[~data['date'].astype(str).str.contains('Date', na=False)]
        data['date'] = pd.to_datetime(data['date'], errors='coerce')
        data = data.dropna(subset=['date']).set_index('date')
        return data.apply(pd.to_numeric, errors='coerce')

    dfs = [parse(CSV_A), parse(CSV_H), parse(CSV_I)]
    price_df = pd.concat(dfs, axis=1)
    price_df = price_df.loc[:, ~price_df.columns.duplicated()].sort_index()
    print(f"    覆盖 {price_df.shape[1]} 只股票/指数，"
          f"{price_df.index[0].date()} ~ {price_df.index[-1].date()}")
    return price_df


# ═══════════════════════════════════════════════════════════════════════════════
# 回测引擎
# ═══════════════════════════════════════════════════════════════════════════════

def next_trade_price(series, target_date):
    """从target_date起第一个有效收盘价（日期，价格）"""
    sub = series[series.index >= pd.Timestamp(target_date)].dropna()
    if len(sub) == 0: return None, None
    return sub.index[0], float(sub.iloc[0])


def calc_stock_returns(series, entry_date, hold_days_list):
    """计算多个持仓期的区间收益率"""
    et, ep = next_trade_price(series, entry_date)
    if et is None or ep == 0: return None
    result = {'entry_date': et, 'entry_price': ep}
    for d in hold_days_list:
        xt, xp = next_trade_price(series, et + timedelta(days=d))
        result[f'ret_{d}d'] = round(xp / ep - 1, 6) if xt is not None and xp and xp != 0 else None
    return result


def assign_fcc_group(fcc, top_thresh, bot_thresh):
    if fcc is None: return '数据缺失'
    if fcc >= top_thresh and fcc > 0: return '加仓'
    if fcc <= bot_thresh and fcc < 0: return '减仓'
    return '持平'


def quartile_label(val, q25, q50, q75):
    if val is None: return None
    if val <= q25: return 'Q1低'
    if val <= q50: return 'Q2'
    if val <= q75: return 'Q3'
    return 'Q4高'


def run_backtest(by_q, price_df, top_pct=0.20):
    """
    主回测函数，返回 DataFrame，每行是一只股票在某季度的信号+收益。
    top_pct: 加减仓分组阈值（默认前/后20%）
    """
    quarters = sorted(by_q.keys(), key=sort_q)
    bench_series = price_df.get(BENCHMARK)
    all_recs = []
    prev_codes = set()

    for q in quarters:
        disc = DISCLOSURE_DATES.get(q)
        if disc is None: continue
        disc_ts = pd.Timestamp(disc)
        recs = by_q.get(q, [])
        if not recs: continue

        print(f"    {q}（披露 {disc}，{len(recs)} 只）...")

        # ── 基准收益 ──
        bench_rets = {}
        if bench_series is not None:
            et, ep = next_trade_price(bench_series, disc_ts)
            for d in HOLD_DAYS:
                if et and ep:
                    xt, xp = next_trade_price(bench_series, et + timedelta(days=d))
                    bench_rets[d] = round(xp/ep-1, 6) if xt and xp and xp != 0 else None
                else:
                    bench_rets[d] = None

        # ── 信号阈值 ──
        fcc_vals = sorted([r['fcc'] for r in recs if r['fcc'] is not None], reverse=True)
        n_top = max(1, int(len(fcc_vals) * top_pct))
        top_thresh = fcc_vals[n_top - 1] if fcc_vals else 0
        bot_thresh = fcc_vals[-n_top]     if fcc_vals else 0

        fw_arr  = [r['fw'] for r in recs if r['fw'] is not None]
        fc_arr  = [r['fc'] for r in recs if r['fc'] is not None]
        fw_q = np.percentile(fw_arr, [25,50,75]) if fw_arr else [0,0,0]
        fc_q = np.percentile(fc_arr, [25,50,75]) if fc_arr else [0,0,0]

        cur_codes = {r['code'] for r in recs}

        for r in recs:
            code = r['code']
            ser = price_df.get(code)
            if ser is None: continue
            ret_data = calc_stock_returns(ser, disc_ts, HOLD_DAYS)
            if ret_data is None: continue

            fcc = r['fcc']; fw = r['fw']; fc = r['fc']
            rec = {
                'q': q, 'code': code, 'name': r['name'], 'ind': r['ind'],
                'fcc': fcc, 'fw': fw, 'fc': fc, 'pv': r['pv'],
                'fcc_group': assign_fcc_group(fcc, top_thresh, bot_thresh),
                'fw_group':  quartile_label(fw, *fw_q),
                'fc_group':  quartile_label(fc, *fc_q),
                'is_new': code not in prev_codes,
            }
            for d in HOLD_DAYS:
                ret   = ret_data.get(f'ret_{d}d')
                bench = bench_rets.get(d)
                rec[f'ret_{d}d']   = round(ret*100, 3)   if ret   is not None else None
                rec[f'bench_{d}d'] = round(bench*100, 3) if bench is not None else None
                rec[f'alpha_{d}d'] = round((ret - bench)*100, 3) \
                                     if ret is not None and bench is not None else None
            all_recs.append(rec)

        prev_codes = cur_codes

    return pd.DataFrame(all_recs), quarters


# ═══════════════════════════════════════════════════════════════════════════════
# 统计汇总
# ═══════════════════════════════════════════════════════════════════════════════

def agg(series):
    s = series.dropna()
    if len(s) == 0:
        return {'mean':None,'median':None,'win':None,'n':0,'std':None}
    return {
        'mean':   round(float(s.mean()),   3),
        'median': round(float(s.median()), 3),
        'win':    round(float((s>0).mean())*100, 1),
        'n':      int(len(s)),
        'std':    round(float(s.std()),    3),
    }


def group_agg(df, group_col, alpha_cols):
    out = {}
    for grp, sub in df.groupby(group_col):
        out[str(grp)] = {col: agg(sub[col]) for col in alpha_cols}
    return out


def quarter_group_agg(df, alpha_cols, quarters):
    """每个季度 × 信号组的平均alpha"""
    out = {}
    for q in quarters:
        sub = df[df['q'] == q]
        if len(sub) == 0: continue
        out[q] = {}
        for grp, gsub in sub.groupby('fcc_group'):
            for col in alpha_cols:
                vals = gsub[col].dropna()
                if len(vals) > 0:
                    out[q][f'{grp}|{col}'] = round(float(vals.mean()), 3)
    return out


def build_dataset(df, quarters):
    df_v = df[df['alpha_30d'].notna()].copy()
    alpha_cols = [f'alpha_{d}d' for d in HOLD_DAYS]

    def safe(v):
        if v is None: return None
        if isinstance(v, float) and (v != v): return None
        return v

    # ── 1. 信号组 alpha ──
    fcc_agg = group_agg(df_v, 'fcc_group', alpha_cols)
    fw_agg  = group_agg(df_v[df_v['fw_group'].notna()], 'fw_group', alpha_cols)
    new_agg = group_agg(df_v, 'is_new', alpha_cols)

    # ── 2. 加仓幅度三分位（仅加仓组内部分） ──
    add_df = df_v[df_v['fcc_group'] == '加仓'].copy()
    if len(add_df) >= 9:
        add_df['fcc_q3'] = pd.qcut(add_df['fcc'], q=3, labels=['低加仓','中加仓','高加仓'])
        fcc_size_agg = group_agg(add_df, 'fcc_q3', alpha_cols)
    else:
        fcc_size_agg = {}

    # ── 3. 逐季度：新进入 vs 老持仓 vs 加仓 vs 减仓 ──
    q_detail = {}
    for q in quarters:
        sub = df_v[df_v['q'] == q]
        if len(sub) == 0: continue
        row = {}
        for grp, mask in [('新进入', sub['is_new']==True),
                           ('老持仓', sub['is_new']==False),
                           ('加仓',   sub['fcc_group']=='加仓'),
                           ('减仓',   sub['fcc_group']=='减仓')]:
            s = sub[mask]['alpha_30d'].dropna()
            if len(s) > 0:
                row[grp] = {'mean': round(float(s.mean()),3),
                            'win':  round(float((s>0).mean())*100,1),
                            'n':    int(len(s))}
        bench_s = sub['bench_30d'].dropna()
        row['bench'] = round(float(bench_s.iloc[0]),3) if len(bench_s) else None
        q_detail[q] = row

    # ── 4. 行业：仅加仓组，≥10条 ──
    ind_agg = {}
    for ind, sub in df_v[df_v['fcc_group']=='加仓'].groupby('ind'):
        for col in alpha_cols:
            vals = sub[col].dropna()
            if len(vals) >= 10:
                if ind not in ind_agg: ind_agg[ind] = {}
                ind_agg[ind][col] = round(float(vals.mean()), 3)
                ind_agg[ind]['n'] = int(len(sub))

    # ── 5. 逐季度新进入胜率（折线用） ──
    new_by_q = {}
    for q in quarters:
        sub = df_v[df_v['q'] == q]
        for is_new, label in [(True,'新进入'),(False,'老持仓')]:
            s = sub[sub['is_new']==is_new]['alpha_30d'].dropna()
            if q not in new_by_q: new_by_q[q] = {}
            if len(s) > 0:
                new_by_q[q][label] = round(float(s.mean()), 3)

    # ── 6. 明细表（新进入+加仓组 top 1000） ──
    detail_cols = ['q','code','name','ind','fcc_group','is_new','fcc','fw','fc','pv'] + alpha_cols
    detail_df = df_v[df_v['fcc_group'].isin(['加仓','减仓']) | df_v['is_new']].copy()
    detail_df = detail_df[detail_cols].sort_values('alpha_30d', ascending=False).head(1000)
    detail_recs = [{k: safe(v) for k,v in row.items()} for row in detail_df.to_dict('records')]

    # ── 7. 基准各季度 ──
    bench_by_q = {}
    for q in quarters:
        sub = df_v[df_v['q'] == q]
        for d in HOLD_DAYS:
            vals = sub[f'bench_{d}d'].dropna()
            if len(vals) > 0:
                if q not in bench_by_q: bench_by_q[q] = {}
                bench_by_q[q][f'{d}d'] = round(float(vals.iloc[0]), 3)

    return {
        'quarters':     quarters,
        'hold_days':    HOLD_DAYS,
        'fcc_agg':      fcc_agg,
        'fw_agg':       fw_agg,
        'new_agg':      new_agg,
        'fcc_size_agg': fcc_size_agg,
        'q_detail':     q_detail,
        'ind_agg':      ind_agg,
        'new_by_q':     new_by_q,
        'detail':       detail_recs,
        'bench_by_q':   bench_by_q,
        'gen_time':     datetime.now().strftime('%Y-%m-%d %H:%M'),
        'total_obs':    int(len(df_v)),
        'quarters_covered': f"{quarters[0]}–{quarters[-1]}",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# HTML 模板（策略结论驱动）
# ═══════════════════════════════════════════════════════════════════════════════

HTML = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>基金持仓信号回测 · 策略结论</title>
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
<style>
:root{
  --bg:#f0ece4;--card:#fff;--th:#f7f5f1;--alt:#faf8f5;
  --hdr:#0f1f3d;--bd:#ddd8d0;--bdt:#ccc8bf;
  --t1:#14213d;--t2:#566070;--t3:#96a0ae;
  --acc:#1e50a2;--acc2:#eef3fc;
  --pos:#c0392b;--neg:#166534;
  --sh:0 2px 8px rgba(20,33,61,.09),0 1px 3px rgba(20,33,61,.05);
  --r:8px;
  --ui:'Sora','PingFang SC','Microsoft YaHei',sans-serif;
  --mono:'JetBrains Mono','Courier New',monospace;
}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--t1);font-family:var(--ui);font-size:13px;line-height:1.5}
::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-thumb{background:var(--bd);border-radius:4px}
.hdr{background:var(--hdr);padding:0 28px;height:56px;display:flex;align-items:center;
  justify-content:space-between;position:sticky;top:0;z-index:50;box-shadow:0 1px 0 rgba(255,255,255,.08)}
.brand{display:flex;align-items:center;gap:12px}
.icon{width:34px;height:34px;background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.18);
  border-radius:8px;display:flex;align-items:center;justify-content:center;flex-shrink:0}
.hdr h1{font-size:15px;font-weight:700;color:#fff}
.hdr-sub{font-size:9px;color:rgba(255,255,255,.38);letter-spacing:1.2px;text-transform:uppercase}
.hdr-meta{color:rgba(255,255,255,.35);font-size:11px;font-family:var(--mono)}
.tabs{background:#fff;padding:0 28px;display:flex;border-bottom:1px solid var(--bd);
  position:sticky;top:56px;z-index:40;box-shadow:0 1px 3px rgba(20,33,61,.07)}
.tab{padding:14px 16px;cursor:pointer;border:none;background:none;color:var(--t3);
  font-size:12px;font-family:var(--ui);font-weight:500;border-bottom:2px solid transparent;
  transition:all .18s;white-space:nowrap;margin-bottom:-1px}
.tab:hover{color:var(--t2)}
.tab.on{color:var(--acc);border-bottom-color:var(--acc);font-weight:700}
.pane{display:none;padding:24px 28px;min-height:calc(100vh - 109px)}
.pane.on{display:block}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px}
.mb{margin-bottom:14px}
.ctrl{display:flex;gap:8px;align-items:center;margin-bottom:18px;flex-wrap:wrap}
.lbl{color:var(--t3);font-size:10px;font-weight:600;letter-spacing:.6px;text-transform:uppercase;white-space:nowrap}
select,input[type=text]{background:var(--card);border:1px solid var(--bd);color:var(--t1);
  padding:5px 10px;border-radius:6px;font-size:12px;font-family:var(--ui);outline:none;height:30px}
input[type=text]{width:200px}
select:focus,input:focus{border-color:var(--acc);box-shadow:0 0 0 3px rgba(30,80,162,.1)}
.cards{display:flex;gap:10px;margin-bottom:18px;flex-wrap:wrap}
.card{background:var(--card);border:1px solid var(--bd);border-radius:var(--r);
  padding:14px 18px;flex:1;min-width:140px;max-width:210px;box-shadow:var(--sh)}
.card-lbl{font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:.7px}
.card-val{font-size:22px;font-weight:700;margin-top:5px;font-family:var(--mono);letter-spacing:-.5px;line-height:1}
.card-sub{font-size:11px;color:var(--t3);margin-top:4px}
.pos{color:var(--pos)}.neg{color:var(--neg)}.neu{color:var(--t3)}
.box{background:var(--card);border:1px solid var(--bd);border-radius:var(--r);padding:16px 18px;box-shadow:var(--sh)}
.box-title{font-size:9px;color:var(--t3);margin-bottom:10px;font-weight:700;text-transform:uppercase;letter-spacing:.7px}
.strat{background:var(--card);border:1px solid var(--bd);border-radius:var(--r);
  padding:20px 22px;box-shadow:var(--sh);margin-bottom:14px}
.strat-head{display:flex;align-items:flex-start;gap:14px;margin-bottom:14px}
.strat-num{width:36px;height:36px;border-radius:9px;background:var(--acc);color:#fff;
  font-size:16px;font-weight:700;display:flex;align-items:center;justify-content:center;flex-shrink:0}
.strat-title{font-size:16px;font-weight:700;color:var(--t1);margin-bottom:4px}
.strat-desc{font-size:13px;color:var(--t2);line-height:1.7}
.strat-nums{display:flex;gap:16px;margin:14px 0;flex-wrap:wrap}
.sn{text-align:center;padding:10px 14px;background:var(--th);border-radius:6px;min-width:90px}
.sn-v{font-size:20px;font-weight:700;font-family:var(--mono)}
.sn-l{font-size:10px;color:var(--t3);margin-top:3px}
.action{background:var(--acc2);border:1px solid #c5d8f5;border-radius:6px;
  padding:10px 14px;font-size:12px;color:var(--acc);line-height:1.7;margin-top:10px}
.action b{font-weight:700}
.warn{background:#fffbeb;border:1px solid #fcd34d;border-radius:6px;
  padding:10px 14px;font-size:12px;color:#92400e;line-height:1.7;margin-top:8px}
.warn b{font-weight:700}
.tbl-wrap{overflow-x:auto;border-radius:var(--r);border:1px solid var(--bd);box-shadow:var(--sh);background:var(--card)}
table{width:100%;border-collapse:collapse}
thead{position:sticky;top:0;z-index:10}
th{background:var(--th);color:var(--t3);font-weight:700;font-size:10px;text-transform:uppercase;
  letter-spacing:.5px;padding:10px 12px;text-align:right;white-space:nowrap;border-bottom:1px solid var(--bdt)}
th.l,td.l{text-align:left}
td{padding:8px 12px;border-top:1px solid #ede9e3;text-align:right;white-space:nowrap;
  font-size:12px;font-family:var(--mono);color:var(--t1)}
td.l{font-family:var(--ui);font-size:13px}
tr:nth-child(even) td{background:var(--alt)}
tr:hover td{background:var(--acc2)!important}
.code{color:var(--acc);font-size:11px}
.tag{display:inline-block;background:var(--acc2);color:var(--acc);padding:2px 7px;
  border-radius:4px;font-size:11px;font-weight:600;font-family:var(--ui)}
.tag-add{background:#fdf2f1;color:var(--pos)}
.tag-red{background:#f0f8f3;color:var(--neg)}
.tag-new{background:#fef3c7;color:#92400e}
.divider{border:none;border-top:1px solid var(--bd);margin:22px 0}
.hint{color:var(--t3);font-size:11px;line-height:1.7;margin-top:10px}
.hint b{color:var(--t2)}
</style>
</head>
<body>
<div class="hdr">
  <div class="brand">
    <div class="icon"><svg width="16" height="16" fill="none"><path d="M1 15h4V7H1v8zm5-11h4v11H6V4zm5-3h4v14h-4V1z" fill="white" opacity=".9"/></svg></div>
    <div><h1>基金持仓信号回测 · 策略结论</h1><div class="hdr-sub">Fund Signal Backtest · vs CSI300</div></div>
  </div>
  <span class="hdr-meta" id="meta"></span>
</div>
<div class="tabs">
  <button class="tab on"  onclick="go('t1')">执行摘要</button>
  <button class="tab" onclick="go('t2')">策略一：首次建仓信号</button>
  <button class="tab" onclick="go('t3')">策略二：集中度逆向</button>
  <button class="tab" onclick="go('t4')">策略三：加仓幅度逆向</button>
  <button class="tab" onclick="go('t5')">行业分化</button>
  <button class="tab" onclick="go('t6')">数据明细</button>
</div>

<!-- TAB 1 -->
<div id="t1" class="pane on">
  <div class="cards" id="t1-cards"></div>
  <div class="g2">
    <div class="box"><div class="box-title">五大信号 · 30日超额收益对比</div><div id="c-summary" style="height:300px"></div></div>
    <div class="box"><div class="box-title">信号胜率 · 持仓期敏感性</div><div id="c-winrate" style="height:300px"></div></div>
  </div>
  <div class="strat" style="border-left:4px solid var(--acc)">
    <div style="font-size:13px;font-weight:700;margin-bottom:10px">回测说明</div>
    <div style="font-size:12px;color:var(--t2);line-height:1.8">
      <b>数据区间</b>：<span id="t1-range"></span> · <b>有效观测</b>：<span id="t1-obs"></span> 条 · <b>基准</b>：沪深300（000300.SH）<br>
      <b>信号生效日</b>：季报披露日（约季末次月20日，如1Q→4月20日，2Q→7月20日）<br>
      <b>超额收益定义</b>：个股区间收益 − 同期沪深300收益，按自然日计算持仓期<br>
      <b>加减仓分组</b>：持有基金数变动（fcc）按季度内排名前20%为加仓组，后20%为减仓组
    </div>
  </div>
  <hr class="divider">
  <div style="font-size:13px;font-weight:700;margin-bottom:14px">使用局限性</div>
  <div class="g2">
    <div class="warn"><b>均值 vs 中位数差异大</b>：加仓组30日均值+2.69%，但中位数仅+0.30%。分布右偏，少数大涨股拉高均值，胜率仅51%。</div>
    <div class="warn"><b>回测期含牛市行情</b>：2024Q4–2025Q2为强牛市阶段，整体α偏高，信号在熊市环境的表现尚待验证。</div>
    <div class="warn"><b>披露日存在误差</b>：统一使用每季末次月20日，实际各基金披露时间有±5日偏差。</div>
    <div class="warn"><b>H股匹配率约88%</b>：部分港股代码格式差异导致约12%持仓无法匹配收盘价。</div>
  </div>
</div>

<!-- TAB 2 -->
<div id="t2" class="pane">
  <div class="strat">
    <div class="strat-head">
      <div class="strat-num">1</div>
      <div style="flex:1">
        <div class="strat-title">首次建仓信号 — 最强、最稳定的信号</div>
        <div class="strat-desc">本季度首次出现在公募重仓名单的个股（新进入），披露后30日超额收益均值比老持仓股高约<b>+2pp</b>，胜率高约<b>9pp</b>（59.8% vs 51.0%）。这说明基金新建仓是一个市场尚未充分定价的强看好信号——基金研究员花费大量时间调研后首次建仓，往往领先市场共识。</div>
        <div class="strat-nums" id="t2-nums"></div>
        <div class="action"><b>操作建议：</b>每次季报披露后，优先筛选「本季新进入」股票池，结合行业判断选取科技/工业类标的，30日持仓胜率接近60%。可作为初步筛选门槛，再叠加基本面研究验证。注意首季度（1Q24）所有股票均视为新进入，数据可能有偏。</div>
      </div>
    </div>
  </div>
  <div class="ctrl">
    <span class="lbl">持仓期</span>
    <select id="t2-hold" onchange="renderT2()"><option value="30">30日</option><option value="60">60日</option><option value="90">90日</option></select>
  </div>
  <div class="g2">
    <div class="box"><div class="box-title">新进入 vs 老持仓 · 各持仓期超额收益均值</div><div id="c-new-bar" style="height:280px"></div></div>
    <div class="box"><div class="box-title">逐季度 · 新进入 vs 老持仓 · 30日超额收益</div><div id="c-new-q" style="height:280px"></div></div>
  </div>
  <p class="hint">注：「新进入」定义为本季度重仓名单出现、上季度未出现的股票。首季度（1Q24）所有股票均计为新进入，会拉高该季度新进入均值。</p>
</div>

<!-- TAB 3 -->
<div id="t3" class="pane">
  <div class="strat">
    <div class="strat-head">
      <div class="strat-num" style="background:#7c3aed">2</div>
      <div style="flex:1">
        <div class="strat-title">低集中度逆向策略 — 别只盯着顶级重仓股</div>
        <div class="strat-desc">反直觉发现：占基金持仓比重（fw）<b>最低的Q1组</b>，30日超额+4.00%（胜率59.4%），远优于<b>最高的Q4组</b>（+0.98%，胜率46.1%）。高度集中的明星股已被市场充分定价、持仓拥挤度高；排名靠后的「边缘重仓股」存在更大的预期差空间。随着持仓期延长，低集中度优势仍然保持，说明该信号具有一定的持续性。</div>
        <div class="strat-nums" id="t3-nums"></div>
        <div class="action"><b>操作建议：</b>不要只盯每季度前10大龙头股。重点关注fw排名50–150区间的「安静重仓股」，这类股票有基金认可但关注度低，信息不对称更大。适合中小盘成长股筛选场景。</div>
        <div class="warn"><b>注意：</b>低集中度股样本较分散，流动性相对较弱，不适合大资金直接应用。建议与基本面分析结合使用。</div>
      </div>
    </div>
  </div>
  <div class="ctrl">
    <span class="lbl">持仓期</span>
    <select id="t3-hold" onchange="renderT3()"><option value="30">30日</option><option value="60">60日</option><option value="90">90日</option></select>
  </div>
  <div class="g2">
    <div class="box"><div class="box-title">fw分位数组 · 超额收益均值（Q1低→Q4高）</div><div id="c-fw-bar" style="height:280px"></div></div>
    <div class="box"><div class="box-title">各分位数组 · 持仓期敏感性</div><div id="c-fw-line" style="height:280px"></div></div>
  </div>
  <div class="tbl-wrap">
    <table><thead><tr>
      <th class="l">分位组</th><th>样本数</th>
      <th>30日α均值</th><th>30日α中位</th><th>30日胜率</th>
      <th>60日α均值</th><th>60日α中位</th><th>60日胜率</th>
      <th>90日α均值</th><th>90日α中位</th><th>90日胜率</th>
    </tr></thead><tbody id="t3-body"></tbody></table>
  </div>
</div>

<!-- TAB 4 -->
<div id="t4" class="pane">
  <div class="strat">
    <div class="strat-head">
      <div class="strat-num" style="background:#0891b2">3</div>
      <div style="flex:1">
        <div class="strat-title">加仓幅度逆向 — 小幅加仓比大幅加仓更有价值</div>
        <div class="strat-desc">在所有加仓股票（fcc&gt;0）内部，按加仓幅度三分位分组：<b>低加仓组</b>（刚开始悄悄布局）30日超额+4.36%（胜率57.1%），<b>高加仓组</b>（大幅买入）仅+1.39%（胜率47.1%）。原因在于：大幅加仓往往是市场热点驱动的追涨行为，股价已提前反映；小幅加仓是基金悄悄建仓，市场尚未定价，预期差最大。</div>
        <div class="strat-nums" id="t4-nums"></div>
        <div class="action"><b>操作建议：</b>筛选条件设为「fcc介于1–5家的加仓股」，而非fcc变化最大的票。「基金增减数量爆表」的股票可能存在追高风险，应谨慎对待。低加仓股往往是基金研究员刚开始覆盖的新票，媒体关注度低，是预期差最大的来源。</div>
      </div>
    </div>
  </div>
  <div class="g2 mb">
    <div class="box"><div class="box-title">加仓组内部 · 按加仓幅度三分位 · 30/60/90日超额</div><div id="c-size-bar" style="height:300px"></div></div>
    <div class="box"><div class="box-title">加仓 vs 减仓 · 多空超额差值（各持仓期）</div><div id="c-ls-line" style="height:300px"></div></div>
  </div>
  <div class="strat" style="border-left:4px solid #0891b2;background:var(--alt)">
    <div style="font-size:13px;font-weight:600;margin-bottom:8px">附：加减仓多空策略有效性</div>
    <div style="font-size:12px;color:var(--t2);line-height:1.8">
      整体加仓组超额均值（+2.69%）高于减仓组（+1.41%），30日多空差约 <b id="t4-ls30">—</b>，90日扩大至约 <b id="t4-ls90">—</b>。
      信号有效但幅度有限（胜率仅差约3pp），单一维度不足以支撑高胜率策略，建议与集中度、新进入等信号叠加使用。
    </div>
  </div>
</div>

<!-- TAB 5 -->
<div id="t5" class="pane">
  <div class="strat" style="border-left:4px solid #d97706">
    <div class="strat-head">
      <div class="strat-num" style="background:#d97706">4</div>
      <div style="flex:1">
        <div class="strat-title">行业过滤是信号质量的关键</div>
        <div class="strat-desc">同样是「加仓信号」，不同行业后续超额收益差异巨大：通信（+8.95%）、电力设备（+5.04%）效果显著；食品饮料（−1.53%）、社会服务（−1.42%）基本无效甚至反向。消费类行业基金加仓往往出于防御性配置需求，而非进攻性看好，信号预测力很弱。</div>
        <div class="action"><b>操作建议：</b>加仓信号在科技（通信、电子）、工业（电力设备、建筑材料）板块有效，可信度高；在消费（食品饮料、社会服务、农林牧渔）板块应谨慎，加仓信号的预测力接近随机。将「行业过滤」作为策略一/二/三的前置条件，可显著提升胜率。</div>
      </div>
    </div>
  </div>
  <div class="ctrl">
    <span class="lbl">持仓期</span>
    <select id="t5-hold" onchange="renderT5()"><option value="30">30日</option><option value="60">60日</option><option value="90">90日</option></select>
  </div>
  <div class="box mb"><div class="box-title">各行业加仓信号后续超额收益均值（降序，n≥10）</div><div id="c-ind" style="height:480px"></div></div>
  <div class="g2">
    <div class="box">
      <div class="box-title" style="color:var(--pos)">▲ 信号有效行业 Top 5</div>
      <div id="ind-top"></div>
    </div>
    <div class="box">
      <div class="box-title" style="color:var(--neg)">▼ 信号无效行业 Bottom 5</div>
      <div id="ind-bot"></div>
    </div>
  </div>
</div>

<!-- TAB 6 -->
<div id="t6" class="pane">
  <div class="ctrl">
    <span class="lbl">信号</span>
    <select id="t6-grp" onchange="renderT6()"><option value="">全部</option><option value="加仓">加仓</option><option value="减仓">减仓</option></select>
    <span class="lbl">新进入</span>
    <select id="t6-new" onchange="renderT6()"><option value="">全部</option><option value="true">仅新进入</option><option value="false">仅老持仓</option></select>
    <span class="lbl">季度</span>
    <select id="t6-q" onchange="renderT6()"><option value="">全部</option></select>
    <span class="lbl">排序</span>
    <select id="t6-sort" onchange="renderT6()">
      <option value="alpha_30d_d">30日α↓</option><option value="alpha_30d_a">30日α↑</option>
      <option value="alpha_60d_d">60日α↓</option><option value="fcc_d">加仓最多↓</option><option value="fcc_a">减仓最多↑</option>
    </select>
    <input type="text" id="t6-kw" placeholder="代码/名称" oninput="renderT6()">
  </div>
  <div class="tbl-wrap">
    <table><thead><tr>
      <th class="l">季度</th><th class="l">代码</th><th class="l">名称</th><th class="l">行业</th>
      <th class="l">信号</th><th class="l">新进入</th>
      <th>基金增减</th><th>占基金持仓</th><th>持有基金数</th>
      <th>30日α</th><th>60日α</th><th>90日α</th>
    </tr></thead><tbody id="t6-body"></tbody></table>
  </div>
  <p class="hint" id="t6-cnt"></p>
</div>

<script>
const D = __DATA__;
const FMT = v => v==null?'—':(v>0?'+':'')+v.toFixed(2)+'%';
const CLS = v => v==null?'neu':v>0?'pos':v<0?'neg':'neu';
const AX = {
  axisLine:{lineStyle:{color:'#ccc8bf'}},
  axisLabel:{color:'#96a0ae',fontSize:10,fontFamily:"'JetBrains Mono',monospace"},
  splitLine:{lineStyle:{color:'#e8e3db',type:'dashed'}},
};
const TT = {backgroundColor:'#fff',borderColor:'#ddd8d0',borderWidth:1,textStyle:{color:'#14213d',fontSize:11}};

function mc(id){const el=document.getElementById(id);if(!el)return null;let c=echarts.getInstanceByDom(el);if(c)c.dispose();return echarts.init(el,null,{renderer:'canvas'});}
function go(id){
  document.querySelectorAll('.pane').forEach(p=>p.classList.remove('on'));
  document.querySelectorAll('.tab').forEach(b=>b.classList.remove('on'));
  document.getElementById(id).classList.add('on');
  const idx=['t1','t2','t3','t4','t5','t6'].indexOf(id);
  document.querySelectorAll('.tab')[idx].classList.add('on');
  ({t2:renderT2,t3:renderT3,t4:renderT4,t5:renderT5,t6:renderT6}[id]||function(){})();
}

function renderT1(){
  const fa=D.fcc_agg,na=D.new_agg,fw=D.fw_agg,sz=D.fcc_size_agg;
  const add30=fa['加仓']?.alpha_30d,red30=fa['减仓']?.alpha_30d;
  const new30=na['true']?.alpha_30d,old30=na['false']?.alpha_30d;
  const fwQ1=fw['Q1低']?.alpha_30d,fwQ4=fw['Q4高']?.alpha_30d;
  const szLow=sz['低加仓']?.alpha_30d;
  const cards=[
    {lbl:'① 新进入信号 30日α',val:new30?.mean,sub:`胜率 ${new30?.win}%  老持仓: ${FMT(old30?.mean)}`,c:'#1e50a2'},
    {lbl:'② 低集中度(Q1) 30日α',val:fwQ1?.mean,sub:`胜率 ${fwQ1?.win}%  高集中Q4: ${FMT(fwQ4?.mean)}`,c:'#7c3aed'},
    {lbl:'③ 低加仓幅度 30日α',val:szLow?.mean,sub:`胜率 ${szLow?.win}%  高加仓: ${FMT(sz['高加仓']?.alpha_30d?.mean)}`,c:'#0891b2'},
    {lbl:'④ 加−减 多空差(30日)',val:add30&&red30?+(add30.mean-red30.mean).toFixed(3):null,sub:`加仓${FMT(add30?.mean)} 减仓${FMT(red30?.mean)}`,c:'#c0392b'},
    {lbl:'⑤ 通信行业加仓α',val:D.ind_agg['通信']?.alpha_30d,sub:`n=${D.ind_agg['通信']?.n||0}`,c:'#d97706'},
  ];
  document.getElementById('t1-cards').innerHTML=cards.map(c=>`<div class="card">
    <div class="card-lbl" style="color:${c.c}">${c.lbl}</div>
    <div class="card-val ${CLS(c.val)}">${FMT(c.val)}</div>
    <div class="card-sub">${c.sub}</div></div>`).join('');
  document.getElementById('t1-range').textContent=D.quarters_covered;
  document.getElementById('t1-obs').textContent=D.total_obs.toLocaleString();
  const groups=[
    {n:'新进入',v:new30?.mean,c:'#1e50a2'},{n:'低集中(Q1)',v:fwQ1?.mean,c:'#7c3aed'},
    {n:'低加仓幅度',v:szLow?.mean,c:'#0891b2'},{n:'加仓整体',v:add30?.mean,c:'#d97706'},
    {n:'持平整体',v:fa['持平']?.alpha_30d?.mean,c:'#96a0ae'},{n:'减仓整体',v:red30?.mean,c:'#166534'},
    {n:'老持仓',v:old30?.mean,c:'#ccc8bf'},
  ];
  const cS=mc('c-summary');
  if(cS)cS.setOption({backgroundColor:'transparent',
    grid:{left:90,right:60,top:10,bottom:20},
    xAxis:{type:'value',...AX,axisLine:{show:false},axisLabel:{formatter:v=>v.toFixed(1)+'%'}},
    yAxis:{type:'category',data:groups.map(g=>g.n),inverse:false,axisLine:{show:false},axisTick:{show:false},
      axisLabel:{color:'#566070',fontSize:11,fontFamily:"'Sora',sans-serif"}},
    tooltip:{...TT,trigger:'axis',formatter:p=>`${p[0].name}: ${FMT(p[0].value)}`},
    series:[{type:'bar',data:groups.map(g=>({value:g.v,itemStyle:{color:g.c,borderRadius:[0,4,4,0]}})),
      barMaxWidth:20,label:{show:true,position:'right',color:'#96a0ae',fontSize:10,formatter:p=>FMT(p.value)}}]
  });
  const wrGrps=['新进入','加仓','减仓'];
  const wrC={'新进入':'#1e50a2','加仓':'#c0392b','减仓':'#166534'};
  const cW=mc('c-winrate');
  if(cW)cW.setOption({backgroundColor:'transparent',
    grid:{left:55,right:90,top:20,bottom:30},
    xAxis:{type:'category',data:D.hold_days.map(d=>d+'日'),...AX,axisTick:{show:false}},
    yAxis:{type:'value',min:40,max:70,...AX,axisLine:{show:false},axisLabel:{formatter:v=>v+'%'}},
    tooltip:{...TT,trigger:'axis'},
    legend:{orient:'vertical',right:0,top:'middle',textStyle:{color:'#566070',fontSize:11}},
    series:wrGrps.map(g=>({name:g,type:'line',smooth:.3,
      data:D.hold_days.map(d=>(g==='新进入'?na['true']:fa[g])?.[`alpha_${d}d`]?.win??null),
      itemStyle:{color:wrC[g]},lineStyle:{color:wrC[g],width:2.5},symbolSize:8}))
  });
}

function renderT2(){
  const hold=document.getElementById('t2-hold').value;
  const col=`alpha_${hold}d`;
  const na=D.new_agg;
  const nv=na['true']?.[col],ov=na['false']?.[col];
  document.getElementById('t2-nums').innerHTML=[
    {v:nv?.mean,l:`新进入 ${hold}日α`,f:true},{v:nv?.win,l:'新进入胜率',suf:'%'},
    {v:ov?.mean,l:`老持仓 ${hold}日α`,f:true},{v:ov?.win,l:'老持仓胜率',suf:'%'},
    {v:nv&&ov?+(nv.mean-ov.mean).toFixed(2):null,l:'超额差值',f:true},
  ].map(s=>`<div class="sn"><div class="sn-v ${s.f?CLS(s.v):''}">${s.v!=null?(s.f?FMT(s.v):s.v+(s.suf||'')):'—'}</div><div class="sn-l">${s.l}</div></div>`).join('');
  const cNb=mc('c-new-bar');
  if(cNb)cNb.setOption({backgroundColor:'transparent',
    grid:{left:55,right:90,top:20,bottom:30},
    xAxis:{type:'category',data:D.hold_days.map(d=>d+'日'),...AX,axisTick:{show:false}},
    yAxis:{type:'value',...AX,axisLine:{show:false},axisLabel:{formatter:v=>v.toFixed(1)+'%'}},
    tooltip:{...TT,trigger:'axis'},
    legend:{orient:'vertical',right:0,top:'middle',textStyle:{color:'#566070',fontSize:11}},
    series:[
      {name:'新进入',type:'bar',data:D.hold_days.map(d=>na['true']?.[`alpha_${d}d`]?.mean??null),
       itemStyle:{color:'#1e50a2',borderRadius:[3,3,0,0]},barGap:'10%',barMaxWidth:36,
       label:{show:true,position:'top',color:'#96a0ae',fontSize:10,formatter:p=>FMT(p.value)}},
      {name:'老持仓',type:'bar',data:D.hold_days.map(d=>na['false']?.[`alpha_${d}d`]?.mean??null),
       itemStyle:{color:'#ccc8bf',borderRadius:[3,3,0,0]},barMaxWidth:36,
       label:{show:true,position:'top',color:'#96a0ae',fontSize:10,formatter:p=>FMT(p.value)}},
    ]
  });
  const qs=D.quarters;
  const cNq=mc('c-new-q');
  if(cNq)cNq.setOption({backgroundColor:'transparent',
    grid:{left:55,right:80,top:20,bottom:30},
    xAxis:{type:'category',data:qs,...AX,axisTick:{show:false},axisLabel:{color:'#566070',fontSize:11}},
    yAxis:{type:'value',...AX,axisLine:{show:false},axisLabel:{formatter:v=>v.toFixed(0)+'%'}},
    tooltip:{...TT,trigger:'axis'},
    legend:{right:0,top:0,textStyle:{color:'#566070',fontSize:11}},
    series:[
      {name:'新进入',type:'line',smooth:.3,data:qs.map(q=>D.new_by_q[q]?.['新进入']??null),
       itemStyle:{color:'#1e50a2'},lineStyle:{color:'#1e50a2',width:2.5},symbolSize:7},
      {name:'老持仓',type:'line',smooth:.3,data:qs.map(q=>D.new_by_q[q]?.['老持仓']??null),
       itemStyle:{color:'#96a0ae'},lineStyle:{color:'#96a0ae',width:2,type:'dashed'},symbolSize:6},
    ]
  });
}

function renderT3(){
  const hold=document.getElementById('t3-hold').value;
  const col=`alpha_${hold}d`;
  const fw=D.fw_agg;
  const grps=['Q1低','Q2','Q3','Q4高'];
  const colors=['#4f46e5','#1e50a2','#0891b2','#d97706'];
  const vG=grps.filter(g=>fw[g]);
  document.getElementById('t3-nums').innerHTML=[
    {v:fw['Q1低']?.[col]?.mean,l:`Q1(低) ${hold}日α`,f:true},{v:fw['Q1低']?.[col]?.win,l:'Q1胜率',suf:'%'},
    {v:fw['Q4高']?.[col]?.mean,l:`Q4(高) ${hold}日α`,f:true},{v:fw['Q4高']?.[col]?.win,l:'Q4胜率',suf:'%'},
    {v:fw['Q1低']&&fw['Q4高']?+(fw['Q1低'][col].mean-fw['Q4高'][col].mean).toFixed(2):null,l:'低-高 差值',f:true},
  ].map(s=>`<div class="sn"><div class="sn-v ${s.f?CLS(s.v):''}">${s.v!=null?(s.f?FMT(s.v):s.v+(s.suf||'')):'—'}</div><div class="sn-l">${s.l}</div></div>`).join('');
  const cFb=mc('c-fw-bar');
  if(cFb)cFb.setOption({backgroundColor:'transparent',
    grid:{left:60,right:20,top:20,bottom:30},
    xAxis:{type:'category',data:vG,...AX,axisTick:{show:false},axisLabel:{color:'#566070',fontSize:12}},
    yAxis:{type:'value',...AX,axisLine:{show:false},axisLabel:{formatter:v=>v.toFixed(1)+'%'}},
    tooltip:{...TT,trigger:'axis',formatter:p=>`${p[0].name}<br>均值:${FMT(p[0].value)}<br>胜率:${fw[p[0].name]?.[col]?.win}%<br>n:${fw[p[0].name]?.[col]?.n}`},
    series:[{type:'bar',data:vG.map((g,i)=>({value:fw[g]?.[col]?.mean??null,itemStyle:{color:colors[i],borderRadius:[4,4,0,0]}})),
      barMaxWidth:60,label:{show:true,position:'top',color:'#96a0ae',fontSize:11,formatter:p=>FMT(p.value)}}]
  });
  const cFl=mc('c-fw-line');
  if(cFl)cFl.setOption({backgroundColor:'transparent',
    grid:{left:60,right:80,top:20,bottom:30},
    xAxis:{type:'category',data:D.hold_days.map(d=>d+'日'),...AX,axisTick:{show:false}},
    yAxis:{type:'value',...AX,axisLine:{show:false},axisLabel:{formatter:v=>v.toFixed(1)+'%'}},
    tooltip:{...TT,trigger:'axis'},
    legend:{right:0,top:0,textStyle:{color:'#566070',fontSize:11}},
    series:vG.map((g,i)=>({name:g,type:'line',smooth:.3,
      data:D.hold_days.map(d=>fw[g]?.[`alpha_${d}d`]?.mean??null),
      itemStyle:{color:colors[i]},lineStyle:{color:colors[i],width:2.5},symbolSize:7}))
  });
  document.getElementById('t3-body').innerHTML=vG.map((g,i)=>{
    const r=fw[g]||{};
    return `<tr><td class="l"><span class="tag" style="background:${colors[i]}22;color:${colors[i]}">${g}</span></td>
      <td>${r.alpha_30d?.n??'—'}</td>
      ${D.hold_days.map(d=>`<td class="${CLS(r[`alpha_${d}d`]?.mean)}">${FMT(r[`alpha_${d}d`]?.mean)}</td>
        <td class="${CLS(r[`alpha_${d}d`]?.median)}">${FMT(r[`alpha_${d}d`]?.median)}</td>
        <td>${r[`alpha_${d}d`]?.win??'—'}%</td>`).join('')}</tr>`;
  }).join('');
}

function renderT4(){
  const sz=D.fcc_size_agg,fa=D.fcc_agg;
  const sG=['低加仓','中加仓','高加仓'];
  const sC=['#1e50a2','#d97706','#c0392b'];
  document.getElementById('t4-nums').innerHTML=sG.map(g=>[
    {v:sz[g]?.alpha_30d?.mean,l:`${g} 30日α`,f:true},{v:sz[g]?.alpha_30d?.win,l:`${g} 胜率`,suf:'%'}
  ]).flat().map(s=>`<div class="sn"><div class="sn-v ${s.f?CLS(s.v):''}">${s.v!=null?(s.f?FMT(s.v):s.v+'%'):'—'}</div><div class="sn-l">${s.l}</div></div>`).join('');
  const ls30=fa['加仓']&&fa['减仓']?+(fa['加仓'].alpha_30d.mean-fa['减仓'].alpha_30d.mean).toFixed(2)+'%':'—';
  const ls90=fa['加仓']&&fa['减仓']?+(fa['加仓'].alpha_90d.mean-fa['减仓'].alpha_90d.mean).toFixed(2)+'%':'—';
  document.getElementById('t4-ls30').textContent=ls30;
  document.getElementById('t4-ls90').textContent=ls90;
  const cSz=mc('c-size-bar');
  if(cSz)cSz.setOption({backgroundColor:'transparent',
    grid:{left:60,right:90,top:20,bottom:30},
    xAxis:{type:'category',data:D.hold_days.map(d=>d+'日'),...AX,axisTick:{show:false}},
    yAxis:{type:'value',...AX,axisLine:{show:false},axisLabel:{formatter:v=>v.toFixed(1)+'%'}},
    tooltip:{...TT,trigger:'axis'},
    legend:{orient:'vertical',right:0,top:'middle',textStyle:{color:'#566070',fontSize:11}},
    series:sG.map((g,i)=>({name:g,type:'bar',
      data:D.hold_days.map(d=>sz[g]?.[`alpha_${d}d`]?.mean??null),
      itemStyle:{color:sC[i],borderRadius:[3,3,0,0]},
      label:{show:true,position:'top',color:'#96a0ae',fontSize:10,formatter:p=>FMT(p.value)}}))
  });
  const cLs=mc('c-ls-line');
  if(cLs)cLs.setOption({backgroundColor:'transparent',
    grid:{left:60,right:90,top:20,bottom:30},
    xAxis:{type:'category',data:D.hold_days.map(d=>d+'日'),...AX,axisTick:{show:false}},
    yAxis:{type:'value',...AX,axisLine:{show:false},axisLabel:{formatter:v=>v.toFixed(1)+'%'}},
    tooltip:{...TT,trigger:'axis'},
    legend:{orient:'vertical',right:0,top:'middle',textStyle:{color:'#566070',fontSize:11}},
    series:[
      {name:'加仓组',type:'line',smooth:.3,data:D.hold_days.map(d=>fa['加仓']?.[`alpha_${d}d`]?.mean??null),
       itemStyle:{color:'#c0392b'},lineStyle:{color:'#c0392b',width:2.5},symbolSize:8},
      {name:'减仓组',type:'line',smooth:.3,data:D.hold_days.map(d=>fa['减仓']?.[`alpha_${d}d`]?.mean??null),
       itemStyle:{color:'#166534'},lineStyle:{color:'#166534',width:2.5},symbolSize:8},
      {name:'多空差',type:'line',smooth:.3,data:D.hold_days.map(d=>{
        const a=fa['加仓']?.[`alpha_${d}d`]?.mean,r=fa['减仓']?.[`alpha_${d}d`]?.mean;
        return a!=null&&r!=null?+(a-r).toFixed(3):null;}),
       itemStyle:{color:'#d97706'},lineStyle:{color:'#d97706',width:2,type:'dashed'},symbolSize:7}
    ]
  });
}

function renderT5(){
  const hold=document.getElementById('t5-hold').value;
  const col=`alpha_${hold}d`;
  const inds=Object.entries(D.ind_agg).filter(([,v])=>v[col]!=null).sort((a,b)=>(b[1][col]||0)-(a[1][col]||0));
  const names=inds.map(([n])=>n),vals=inds.map(([,v])=>v[col]);
  const h=Math.max(400,names.length*26);
  document.getElementById('c-ind').style.height=h+'px';
  const cI=mc('c-ind');
  if(cI)cI.setOption({backgroundColor:'transparent',
    grid:{left:90,right:60,top:10,bottom:20},
    xAxis:{type:'value',...AX,axisLine:{show:false},axisLabel:{formatter:v=>v.toFixed(1)+'%'}},
    yAxis:{type:'category',data:names,inverse:false,axisLine:{show:false},axisTick:{show:false},
      axisLabel:{color:'#566070',fontSize:11,fontFamily:"'Sora',sans-serif"}},
    tooltip:{...TT,trigger:'axis',formatter:p=>`${p[0].name}: ${FMT(p[0].value)}  n=${D.ind_agg[p[0].name]?.n}`},
    series:[{type:'bar',data:vals.map(v=>({value:v,itemStyle:{color:v>=0?'#c0392b':'#166534',borderRadius:[0,4,4,0]}})),
      barMaxWidth:18,label:{show:true,position:'right',color:'#96a0ae',fontSize:10,formatter:p=>FMT(p.value)}}]
  });
  const row=(ind,v,i)=>`<div style="display:flex;align-items:center;gap:8px;padding:9px 0;border-bottom:1px solid #ede9e3">
    <span style="font-size:12px;font-family:var(--mono);color:var(--t3);width:18px">${i+1}</span>
    <span class="tag">${ind}</span><span style="flex:1"></span>
    <span class="sn-v ${CLS(v)}" style="font-size:14px">${FMT(v)}</span>
    <span style="color:var(--t3);font-size:10px">n=${D.ind_agg[ind]?.n}</span></div>`;
  document.getElementById('ind-top').innerHTML=inds.slice(0,5).map(([n,v],i)=>row(n,v[col],i)).join('');
  document.getElementById('ind-bot').innerHTML=inds.slice(-5).reverse().map(([n,v],i)=>row(n,v[col],i)).join('');
}

function renderT6(){
  const grp=document.getElementById('t6-grp').value;
  const nf=document.getElementById('t6-new').value;
  const q=document.getElementById('t6-q').value;
  const sort=document.getElementById('t6-sort').value;
  const kw=document.getElementById('t6-kw').value.trim().toLowerCase();
  const isAsc=sort.endsWith('_a');
  const sk=sort.slice(0,-2);
  let rows=D.detail;
  if(grp) rows=rows.filter(r=>r.fcc_group===grp);
  if(nf==='true') rows=rows.filter(r=>r.is_new===true);
  if(nf==='false') rows=rows.filter(r=>r.is_new===false);
  if(q) rows=rows.filter(r=>r.q===q);
  if(kw) rows=rows.filter(r=>r.code.toLowerCase().includes(kw)||r.name.includes(kw));
  rows=[...rows].sort((a,b)=>{const av=a[sk]??-Infinity,bv=b[sk]??-Infinity;return isAsc?av-bv:bv-av;});
  document.getElementById('t6-cnt').textContent=`共 ${rows.length} 条，显示前500条`;
  document.getElementById('t6-body').innerHTML=rows.slice(0,500).map(r=>{
    const fg=r.fcc_group==='加仓'?'tag-add':r.fcc_group==='减仓'?'tag-red':'';
    return `<tr>
      <td class="l">${r.q}</td><td class="l code">${r.code}</td><td class="l">${r.name}</td>
      <td class="l"><span class="tag" style="font-size:10px">${r.ind||'—'}</span></td>
      <td class="l"><span class="tag ${fg}">${r.fcc_group}</span></td>
      <td class="l">${r.is_new?'<span class="tag tag-new">新进入</span>':'—'}</td>
      <td class="${r.fcc>0?'pos':r.fcc<0?'neg':''}">${r.fcc!=null?(r.fcc>0?'+':'')+r.fcc:'—'}</td>
      <td>${r.fw!=null?(r.fw*100).toFixed(3)+'%':'—'}</td><td>${r.fc??'—'}</td>
      <td class="${CLS(r.alpha_30d)}">${FMT(r.alpha_30d)}</td>
      <td class="${CLS(r.alpha_60d)}">${FMT(r.alpha_60d)}</td>
      <td class="${CLS(r.alpha_90d)}">${FMT(r.alpha_90d)}</td></tr>`;
  }).join('');
}

(function(){
  document.getElementById('meta').textContent=`${D.quarters_covered}  ·  ${D.total_obs.toLocaleString()} obs  ·  生成 ${D.gen_time}`;
  const t6q=document.getElementById('t6-q');
  D.quarters.forEach(q=>{const o=document.createElement('option');o.value=o.text=q;t6q.appendChild(o);});
  renderT1();
})();
</script>
</body>
</html>
"""


# ═══════════════════════════════════════════════════════════════════════════════
# 主函数
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("=== 基金持仓信号回测 ===")
    by_q = load_holdings()
    price_df = load_prices()

    print("运行回测（top_pct=20%）...")
    df, quarters = run_backtest(by_q, price_df, top_pct=0.20)
    print(f"  有效观测: {len(df)} 条")
    print(f"  价格匹配率: {df['alpha_30d'].notna().sum()}/{len(df)} ({df['alpha_30d'].notna().mean()*100:.1f}%)")

    print("构建数据集...")
    dataset = build_dataset(df, quarters)

    print("生成 HTML...")
    data_json = json.dumps(dataset, ensure_ascii=False, default=str, separators=(",", ":"))
    html = HTML.replace("__DATA__", data_json)
    OUTPUT.write_text(html, encoding="utf-8")
    size_kb = OUTPUT.stat().st_size / 1024
    print(f"  输出: {OUTPUT}  ({size_kb:.0f} KB)")
    print("完成！用浏览器打开 fund_backtest.html 查看回测报告。")


if __name__ == "__main__":
    main()
