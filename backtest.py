#!/usr/bin/env python3
"""
基金重仓回测系统 v1.0
分析机构加减仓行为与后续股价表现的关系

运行: python backtest.py
输出: backtest_report.html
"""

import json
import re
import warnings
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl

warnings.filterwarnings("ignore")

BASE    = Path("workwork/research/策略研究/基金重仓看板")
EXCEL   = BASE / "20260122 基金持仓个股视角.xlsx"
A_CSV   = BASE / "20260427 A股日收盘价.csv"
H_CSV   = BASE / "20260427 H股日收盘价.csv"
IDX_CSV = BASE / "20260427 指数收盘点位.csv"
OUTPUT  = BASE / "backtest_report.html"

HORIZONS = [5, 20, 60]  # 持有期（交易日）

# 季度 → 信息披露日 (T=0)
# 取法定截止日作为保守上限，彻底避免未来信息偏差
DISCLOSURE = {
    "1Q24": pd.Timestamp("2024-04-30"),
    "2Q24": pd.Timestamp("2024-08-31"),
    "3Q24": pd.Timestamp("2024-10-31"),
    "4Q24": pd.Timestamp("2025-02-28"),
    "1Q25": pd.Timestamp("2025-04-30"),
    "2Q25": pd.Timestamp("2025-08-31"),
    "3Q25": pd.Timestamp("2025-10-31"),
    "4Q25": pd.Timestamp("2026-02-28"),
    "1Q26": pd.Timestamp("2026-04-30"),
}

COL = {
    "code": 0, "name": 1, "end_price": 2, "shares": 3, "shares_chg": 4,
    "pos_value": 5, "fund_weight": 6, "float_pct": 7,
    "val_chg": 9, "industry": 11, "fund_count": 23, "fund_chg": 24,
}
COL_OLD = {
    "name": 0, "shares": 1, "float_pct": 2, "shares_chg": 3,
    "chg_pct": 4, "pos_value": 5, "code": 6, "industry": 7,
    "fund_count": 8, "fund_chg": 9,
}
QUARTER_RE = re.compile(r"^([1-4]Q\d{2})\s*个股\s*$")

# ═══════════════════════════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════════════════════════

def pv(v):
    if v is None:
        return None
    try:
        f = float(v)
        return None if f != f else round(f, 6)
    except Exception:
        return None

def sort_qkey(q):
    m = re.match(r"(\d)Q(\d{2})", q)
    return (int(m.group(2)), int(m.group(1))) if m else (0, 0)

def group_stats(vals):
    """核心统计量：均值、标准差、胜率、t检验、中位数"""
    clean = [x for x in vals if x is not None and not (isinstance(x, float) and x != x)]
    n = len(clean)
    if n == 0:
        return {"n": 0, "mean": None, "std": None, "hit": None, "tstat": None, "median": None}
    arr = np.array(clean, dtype=float)
    m = float(np.mean(arr))
    s = float(np.std(arr, ddof=1)) if n > 1 else 0.0
    tstat = round(m / (s / np.sqrt(n)), 2) if s > 0 else None
    return {
        "n": n,
        "mean": round(m, 4),
        "std": round(s, 4),
        "hit": round(float(np.sum(arr > 0) / n), 3),
        "tstat": tstat,
        "median": round(float(np.median(arr)), 4),
    }

def normalize_hk(code):
    """Convert HK code to 4-digit format matching the CSV columns."""
    if code.endswith(".HK"):
        num = code.split(".")[0]
        return num.zfill(4) + ".HK"
    return code

# ═══════════════════════════════════════════════════════════════
# 1. 加载基金持仓
# ═══════════════════════════════════════════════════════════════

def _is_old_fmt(ws):
    first = next(ws.iter_rows(min_row=1, max_row=1, max_col=1, values_only=True), [None])[0]
    return isinstance(first, str) and "名称" in str(first)

def load_fund_holdings():
    print("加载基金持仓...")
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    records, quarters_seen = [], set()

    for sname in wb.sheetnames:
        m = QUARTER_RE.match(sname)
        if not m:
            continue
        q = m.group(1)
        if q in ("3Q23", "4Q23"):
            continue
        quarters_seen.add(q)
        ws = wb[sname]

        if _is_old_fmt(ws):
            for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
                code = row[COL_OLD["code"]] if len(row) > COL_OLD["code"] else None
                if not code or not isinstance(code, str) or "." not in str(code):
                    continue
                go = lambda k: row[COL_OLD[k]] if len(row) > COL_OLD[k] else None
                fp_raw = pv(go("float_pct"))
                records.append({
                    "q": q, "code": str(code).strip(),
                    "name": str(go("name") or "").strip(),
                    "ind": str(go("industry") or "其他").strip() or "其他",
                    "sh": pv(go("shares")), "sch": pv(go("shares_chg")),
                    "pv_val": pv(go("pos_value")), "fw": None, "vc": None,
                    "fp": round(fp_raw / 100, 6) if fp_raw is not None else None,
                    "fc": pv(go("fund_count")), "fcc": pv(go("fund_chg")),
                })
        else:
            col_map = dict(COL)
            hdr = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
            if hdr:
                for i, cell in enumerate(hdr):
                    if cell == "持有基金数":
                        col_map["fund_count"] = i
                    elif cell == "基金增减数量":
                        col_map["fund_chg"] = i
            max_col = max(col_map.values()) + 1
            for row in ws.iter_rows(min_row=3, max_col=max_col, values_only=True):
                code = row[col_map["code"]] if len(row) > col_map["code"] else None
                if not code or not isinstance(code, str) or "." not in str(code):
                    continue
                g = lambda k: row[col_map[k]] if len(row) > col_map[k] else None
                records.append({
                    "q": q, "code": str(code).strip(),
                    "name": str(g("name") or "").strip(),
                    "ind": str(g("industry") or "其他").strip() or "其他",
                    "sh": pv(g("shares")), "sch": pv(g("shares_chg")),
                    "pv_val": pv(g("pos_value")), "fw": pv(g("fund_weight")),
                    "fp": pv(g("float_pct")), "fc": pv(g("fund_count")),
                    "fcc": pv(g("fund_chg")), "vc": pv(g("val_chg")),
                })

    wb.close()
    quarters = sorted(quarters_seen, key=sort_qkey)
    df = pd.DataFrame(records)
    print(f"  季度: {quarters}  记录: {len(df)}")
    return quarters, df

# ═══════════════════════════════════════════════════════════════
# 2. 加载价格数据
# ═══════════════════════════════════════════════════════════════

def load_price_csv(path):
    """Wide-format price CSV: 3 metadata rows, then name row, code row, data rows."""
    raw = pd.read_csv(path, skiprows=3, header=None, encoding="utf-8-sig", low_memory=False)
    codes = raw.iloc[1].tolist()          # row 1 = codes: Date, 000001.SZ, ...
    data = raw.iloc[2:].copy()
    data.columns = codes
    data = data.rename(columns={codes[0]: "Date"})
    data = data.set_index("Date")
    data.index = pd.to_datetime(data.index, errors="coerce")
    data = data[data.index.notna()]
    data = data.apply(pd.to_numeric, errors="coerce")
    return data.sort_index()

# ═══════════════════════════════════════════════════════════════
# 3. 构建回测数据集
# ═══════════════════════════════════════════════════════════════

def build_dataset(quarters, fund_df, a_prices, h_prices, idx_prices):
    print("计算信号与收益率...")

    a_tdays = a_prices.index
    h_tdays = h_prices.index
    idx_tdays = idx_prices.index

    h_cols = set(h_prices.columns)
    a_cols = set(a_prices.columns)

    # 基准指数列
    bench_a_col = "881001.WI" if "881001.WI" in idx_prices.columns else "000300.SH"
    bench_h_col = "HSI.HI" if "HSI.HI" in idx_prices.columns else None

    # 构建前一季度数据查找表
    prev_lookup = {}  # (code, q) -> {fw, fp, fc}
    for _, r in fund_df.iterrows():
        prev_lookup[(r["code"], r["q"])] = {"fw": r["fw"], "fp": r["fp"], "fc": r["fc"]}

    all_rows = []

    for qi, q in enumerate(quarters):
        if q not in DISCLOSURE:
            continue
        t0 = DISCLOSURE[q]
        prev_q = quarters[qi - 1] if qi > 0 else None

        # 在各交易日历中定位 t0（取第一个 >= t0 的交易日）
        a_t0i   = int(a_tdays.searchsorted(t0, side="left"))
        h_t0i   = int(h_tdays.searchsorted(t0, side="left"))
        idx_t0i = int(idx_tdays.searchsorted(t0, side="left"))

        if a_t0i >= len(a_tdays):
            print(f"  警告: {q} 披露日 {t0.date()} 超出价格数据范围，跳过")
            continue

        # 预计算各持有期基准收益（整个季度只算一次）
        bret_a, bret_h = {}, {}
        for h in HORIZONS:
            idx_tni = idx_t0i + h
            if idx_tni < len(idx_tdays):
                p0a = idx_prices[bench_a_col].iloc[idx_t0i]
                pna = idx_prices[bench_a_col].iloc[idx_tni]
                bret_a[h] = float((pna - p0a) / p0a) if (p0a and p0a > 0) else None
                if bench_h_col:
                    p0h = idx_prices[bench_h_col].iloc[idx_t0i]
                    pnh = idx_prices[bench_h_col].iloc[idx_tni]
                    bret_h[h] = float((pnh - p0h) / p0h) if (p0h and p0h > 0) else None
                else:
                    bret_h[h] = None
            else:
                bret_a[h] = bret_h[h] = None

        qdf = fund_df[fund_df["q"] == q]
        covered = 0

        for _, row in qdf.iterrows():
            code  = row["code"]
            is_hk = code.endswith(".HK")

            # 匹配价格列
            if is_hk:
                norm = normalize_hk(code)
                if norm not in h_cols:
                    continue
                price_s = h_prices[norm]
                t0i = h_t0i
                tdays = h_tdays
            else:
                if code not in a_cols:
                    continue
                price_s = a_prices[code]
                t0i = a_t0i
                tdays = a_tdays

            if t0i >= len(tdays):
                continue

            fcc = row["fcc"]
            fw  = row["fw"]
            fp  = row["fp"]
            fc  = row["fc"]

            # 环比变动信号
            fw_delta = fp_delta = fc_delta = None
            if prev_q:
                pr = prev_lookup.get((code, prev_q))
                if pr:
                    if fw is not None and pr["fw"] is not None:
                        fw_delta = round((fw - pr["fw"]) * 100, 4)
                    if fp is not None and pr["fp"] is not None:
                        fp_delta = round((fp - pr["fp"]) * 100, 4)
                    if fc is not None and pr["fc"] is not None:
                        fc_delta = float(fc - pr["fc"])

            # 仓位变动幅度（相对于调整后持股量）：sch / (sh - sch) ≈ 相对增减幅
            sch = row.get("sch")
            sh  = row.get("sh")
            sch_pct = None
            if sch is not None and sh is not None and sh is not None and (sh - sch) > 0:
                sch_pct = round(float(sch / (sh - sch)) * 100, 2)

            # 各持有期收益
            ret_vals = {}
            for h in HORIZONS:
                tni = t0i + h
                if tni >= len(tdays):
                    ret_vals[f"ret{h}"] = ret_vals[f"xret{h}"] = None
                    continue
                try:
                    p0 = float(price_s.iloc[t0i])
                    pn = float(price_s.iloc[tni])
                except Exception:
                    ret_vals[f"ret{h}"] = ret_vals[f"xret{h}"] = None
                    continue

                if p0 != p0 or pn != pn or p0 == 0:  # NaN check
                    ret_vals[f"ret{h}"] = ret_vals[f"xret{h}"] = None
                    continue

                ret = (pn - p0) / p0
                bench = bret_h[h] if is_hk else bret_a[h]
                xret = (ret - bench) if bench is not None else None
                ret_vals[f"ret{h}"]  = round(ret, 4)
                ret_vals[f"xret{h}"] = round(xret, 4) if xret is not None else None

            all_rows.append({
                "q": q, "code": code, "name": row["name"], "ind": row["ind"],
                "is_hk": is_hk,
                "fcc": fcc, "fw": fw, "fp": fp, "fc": fc,
                "fw_delta": fw_delta, "fp_delta": fp_delta, "fc_delta": fc_delta,
                "sch_pct": sch_pct, "pv_val": row.get("pv_val"),
                **ret_vals,
            })
            covered += 1

        print(f"  {q}: {len(qdf)} 持仓 → {covered} 匹配到价格数据  "
              f"基准T+20={bret_a.get(20, 0) and round(bret_a[20]*100,2)}%")

    df = pd.DataFrame(all_rows)
    print(f"  总回测记录: {len(df)}  有效T+20超额收益: {df['xret20'].notna().sum()}")
    return df

# ═══════════════════════════════════════════════════════════════
# 4. 策略分析
# ═══════════════════════════════════════════════════════════════

def quintile_analysis(df, signal_col):
    """按信号列分5组，计算各组各持有期超额收益统计。"""
    valid = df[df[signal_col].notna()].copy()
    if len(valid) < 50:
        return []
    valid["_q"] = pd.qcut(valid[signal_col], q=5, labels=["Q1","Q2","Q3","Q4","Q5"])
    result = []
    for lbl in ["Q1","Q2","Q3","Q4","Q5"]:
        grp = valid[valid["_q"] == lbl]
        entry = {"label": lbl, "n": len(grp)}
        for h in HORIZONS:
            entry[f"s{h}"] = group_stats(grp[f"xret{h}"].tolist())
        result.append(entry)
    return result

def direction_analysis(df):
    """Strategy A: 按 fcc 正负分方向。"""
    valid = df[df["fcc"].notna()]
    groups = {"add": valid[valid["fcc"] > 0],
              "flat": valid[valid["fcc"] == 0],
              "reduce": valid[valid["fcc"] < 0]}
    result = {}
    for name, grp in groups.items():
        entry = {"n": len(grp)}
        for h in HORIZONS:
            entry[f"s{h}"] = group_stats(grp[f"xret{h}"].tolist())
        result[name] = entry
    return result

def composite_analysis(df):
    """Strategy D: fcc排名 + fw_delta排名 组合信号。"""
    valid = df[df["fcc"].notna() & df["fw_delta"].notna()].copy()
    if len(valid) < 50:
        return []
    valid["_combo"] = (valid["fcc"].rank(pct=True) + valid["fw_delta"].rank(pct=True)) / 2
    valid["_g"] = pd.qcut(valid["_combo"], q=5, labels=["C1","C2","C3","C4","C5"])
    result = []
    for lbl in ["C1","C2","C3","C4","C5"]:
        grp = valid[valid["_g"] == lbl]
        entry = {"label": lbl, "n": len(grp)}
        for h in HORIZONS:
            entry[f"s{h}"] = group_stats(grp[f"xret{h}"].tolist())
        result.append(entry)
    return result

def consensus_analysis(df):
    """Strategy E: fcc阈值过滤——多少只基金同时加仓时信号最强。"""
    result = {}
    for thr in [1, 3, 5, 10, 15]:
        grp = df[df["fcc"].notna() & (df["fcc"] >= thr)]
        entry = {"n": len(grp), "thr": thr}
        for h in HORIZONS:
            entry[f"s{h}"] = group_stats(grp[f"xret{h}"].tolist())
        result[str(thr)] = entry
    return result

def new_entrant_analysis(df):
    """Strategy F: 新晋重仓（机构首次大量进入）vs 拥挤减仓（高拥挤度但在减仓）。"""
    valid = df[df["fcc"].notna() & df["fc"].notna() & df["fp"].notna()].copy()
    # 新晋: fcc >= 5 且占流通股不高（< 2%），机构刚开始建仓
    fresh = valid[(valid["fcc"] >= 5) & (valid["fp"] < 0.02)]
    # 拥挤减仓: 流通股占比 > 3% 且 fcc <= -3，机构集中出逃
    crowd_exit = valid[(valid["fp"] > 0.03) & (valid["fcc"] <= -3)]
    # 拥挤加仓: 已高拥挤但继续加，动量还是过度？
    crowd_add = valid[(valid["fp"] > 0.03) & (valid["fcc"] >= 3)]

    result = {}
    for name, grp in [("fresh", fresh), ("crowd_exit", crowd_exit), ("crowd_add", crowd_add)]:
        entry = {"n": len(grp)}
        for h in HORIZONS:
            entry[f"s{h}"] = group_stats(grp[f"xret{h}"].tolist())
        result[name] = entry
    return result

def fw_magnitude_analysis(df):
    """Strategy G: fw_delta绝对值分组——大幅调仓 vs 小幅微调的区别。"""
    valid = df[df["fw_delta"].notna()].copy()
    if len(valid) < 30:
        return {}
    # 加仓方向：fw_delta > 0 分大/中/小
    adds = valid[valid["fw_delta"] > 0].copy()
    if len(adds) > 30:
        adds["_sz"] = pd.qcut(adds["fw_delta"], q=3, labels=["小幅","中幅","大幅"])
        result_add = {}
        for lbl in ["小幅","中幅","大幅"]:
            grp = adds[adds["_sz"] == lbl]
            entry = {"n": len(grp)}
            for h in HORIZONS:
                entry[f"s{h}"] = group_stats(grp[f"xret{h}"].tolist())
            result_add[lbl] = entry
    else:
        result_add = {}
    # 减仓方向：fw_delta < 0 分大/中/小
    reds = valid[valid["fw_delta"] < 0].copy()
    if len(reds) > 30:
        reds["_sz"] = pd.qcut(reds["fw_delta"].abs(), q=3, labels=["小幅","中幅","大幅"])
        result_red = {}
        for lbl in ["小幅","中幅","大幅"]:
            grp = reds[reds["_sz"] == lbl]
            entry = {"n": len(grp)}
            for h in HORIZONS:
                entry[f"s{h}"] = group_stats(grp[f"xret{h}"].tolist())
            result_red[lbl] = entry
    else:
        result_red = {}
    return {"add": result_add, "reduce": result_red}

def industry_analysis(df):
    """各行业的加仓/减仓超额收益差。"""
    result = {}
    for ind, grp in df.groupby("ind"):
        if len(grp) < 8:
            continue
        adds = grp[grp["fcc"] > 0] if "fcc" in grp.columns else pd.DataFrame()
        reds = grp[grp["fcc"] < 0] if "fcc" in grp.columns else pd.DataFrame()
        entry = {"n": len(grp), "n_add": len(adds), "n_red": len(reds)}
        for h in HORIZONS:
            col = f"xret{h}"
            entry[f"add_s{h}"] = group_stats(adds[col].tolist()) if len(adds) else {}
            entry[f"red_s{h}"] = group_stats(reds[col].tolist()) if len(reds) else {}
        result[ind] = entry
    return result

def quarter_analysis(df, quarters):
    """各季度的信号有效性。"""
    result = {}
    for q in quarters:
        grp = df[df["q"] == q]
        adds = grp[grp["fcc"] > 0] if "fcc" in grp.columns else pd.DataFrame()
        reds = grp[grp["fcc"] < 0] if "fcc" in grp.columns else pd.DataFrame()
        entry = {"n": len(grp), "n_add": len(adds), "n_red": len(reds)}
        for h in HORIZONS:
            col = f"xret{h}"
            entry[f"add_s{h}"] = group_stats(adds[col].tolist()) if len(adds) else {}
            entry[f"red_s{h}"] = group_stats(reds[col].tolist()) if len(reds) else {}
        result[q] = entry
    return result

# ═══════════════════════════════════════════════════════════════
# 5. 汇总
# ═══════════════════════════════════════════════════════════════

def build_analysis(quarters, df):
    print("汇总策略分析...")

    # 精简 records 用于前端明细表（只保留关键列）
    rec_cols = ["q","code","name","ind","is_hk","fcc","fw","fp","fc",
                "fw_delta","fp_delta","sch_pct","pv_val",
                "ret5","ret20","ret60","xret5","xret20","xret60"]
    records = []
    for _, r in df[rec_cols].iterrows():
        d = {}
        for c in rec_cols:
            v = r[c]
            if isinstance(v, float) and v != v:
                d[c] = None
            elif isinstance(v, (np.floating, np.integer)):
                d[c] = v.item()
            elif isinstance(v, bool) or isinstance(v, np.bool_):
                d[c] = bool(v)
            else:
                d[c] = v
        records.append(d)

    return {
        "meta": {
            "quarters": quarters,
            "disclosure": {q: str(d.date()) for q, d in DISCLOSURE.items() if q in quarters},
            "n_total": len(df),
            "n_valid20": int(df["xret20"].notna().sum()),
            "bench_a": "万得全A",
            "bench_h": "恒生指数",
            "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        },
        "strat_a": direction_analysis(df),
        "strat_b_fcc": quintile_analysis(df, "fcc"),
        "strat_b_fwd": quintile_analysis(df, "fw_delta"),
        "strat_c": composite_analysis(df),
        "strat_d": consensus_analysis(df),
        "strat_e": new_entrant_analysis(df),
        "strat_f": fw_magnitude_analysis(df),
        "by_ind": industry_analysis(df),
        "by_q": quarter_analysis(df, quarters),
        "records": records,
    }

# ═══════════════════════════════════════════════════════════════
# 6. HTML 模板
# ═══════════════════════════════════════════════════════════════

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>基金重仓回测报告</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
<style>
:root{
  --bg:#f0ece4;--bg-card:#fff;--bg-th:#f7f5f1;--bg-alt:#faf8f5;
  --hdr:#14213d;--text1:#14213d;--text2:#566070;--text3:#96a0ae;
  --acc:#1e50a2;--acc-sub:#eef3fc;
  --pos:#c0392b;--pos-bg:#fdf2f1;--neg:#166534;--neg-bg:#f0f8f3;
  --bd:#ddd8d0;--bd-th:#ccc8bf;
  --sh:0 2px 8px rgba(20,33,61,.09),0 1px 3px rgba(20,33,61,.05);
  --r:8px;--mono:'JetBrains Mono','Courier New',monospace;
  --ui:'Sora','PingFang SC','Microsoft YaHei',sans-serif;
}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--text1);font-family:var(--ui);font-size:13px;line-height:1.5}
::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-thumb{background:var(--bd);border-radius:4px}

.hdr{background:var(--hdr);padding:0 24px;height:52px;display:flex;align-items:center;
  justify-content:space-between;position:sticky;top:0;z-index:50}
.hdr-brand{display:flex;align-items:center;gap:10px}
.hdr h1{font-size:14px;font-weight:700;color:#fff;letter-spacing:.2px}
.hdr-sub{font-size:9px;color:rgba(255,255,255,.38);letter-spacing:1.2px;text-transform:uppercase}
.hdr .meta{color:rgba(255,255,255,.35);font-size:11px;font-family:var(--mono)}

.tabs{background:#fff;padding:0 24px;display:flex;border-bottom:1px solid var(--bd);
  position:sticky;top:52px;z-index:40;overflow-x:auto}
.tb{padding:13px 16px;cursor:pointer;border:none;background:none;color:var(--text3);
  font-size:12px;font-family:var(--ui);font-weight:500;border-bottom:2px solid transparent;
  transition:all .15s;white-space:nowrap;margin-bottom:-1px}
.tb.active{color:var(--acc);border-bottom-color:var(--acc);font-weight:700}

.pane{display:none;padding:20px 24px;min-height:calc(100vh - 101px)}
.pane.active{display:block}

.ctrl{display:flex;gap:8px;align-items:center;margin-bottom:16px;flex-wrap:wrap}
.lbl{color:var(--text3);font-size:10px;font-weight:600;letter-spacing:.6px;text-transform:uppercase}
select,input[type=text]{background:#fff;border:1px solid var(--bd);color:var(--text1);
  padding:5px 10px;border-radius:6px;font-size:12px;font-family:var(--ui);height:30px;outline:none}
select:focus,input:focus{border-color:var(--acc)}

.cards{display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap}
.card{background:#fff;border:1px solid var(--bd);border-radius:var(--r);
  padding:12px 16px;min-width:130px;box-shadow:var(--sh)}
.c-lbl{color:var(--text3);font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:.7px}
.c-val{font-size:20px;font-weight:700;color:var(--text1);margin-top:4px;
  font-family:var(--mono);letter-spacing:-.5px;line-height:1.1}
.c-sub{font-size:11px;color:var(--text3);margin-top:3px}

.box{background:#fff;border:1px solid var(--bd);border-radius:var(--r);
  padding:14px 16px;box-shadow:var(--sh)}
.box-title{font-size:9px;color:var(--text3);margin-bottom:10px;
  font-weight:700;text-transform:uppercase;letter-spacing:.7px}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px}
.g3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px;margin-bottom:14px}
.mb{margin-bottom:14px}

.tw{overflow-x:auto;border-radius:var(--r);border:1px solid var(--bd);
  box-shadow:var(--sh);background:#fff}
table{width:100%;border-collapse:collapse}
thead{position:sticky;top:0;z-index:5}
th{background:var(--bg-th);color:var(--text3);font-weight:700;font-size:10px;
  text-transform:uppercase;letter-spacing:.5px;padding:9px 11px;text-align:right;
  white-space:nowrap;cursor:pointer;border-bottom:1px solid var(--bd-th)}
th.l,td.l{text-align:left}
th:hover{color:var(--acc)}
th.asc::after{content:' ↑';color:var(--acc)}
th.desc::after{content:' ↓';color:var(--acc)}
td{padding:7px 11px;border-top:1px solid #ede9e3;text-align:right;
  white-space:nowrap;font-size:12px;font-family:var(--mono)}
td.l{font-family:var(--ui)}
tr:nth-child(even) td{background:var(--bg-alt)}
tr:hover td{background:var(--acc-sub)!important}
.pos{color:var(--pos);font-weight:600}
.neg{color:var(--neg);font-weight:600}
.neu{color:var(--text3)}
.tag{display:inline-block;background:var(--acc-sub);color:var(--acc);
  padding:2px 7px;border-radius:4px;font-size:11px;font-family:var(--ui);font-weight:600}
.code{color:var(--acc);font-size:11px}
.note{color:var(--text3);font-size:11px;margin-top:8px;padding:8px 12px;
  background:var(--bg-alt);border-radius:6px;border-left:3px solid var(--bd)}
.sig-badge{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600}
.sig-pos{background:var(--pos-bg);color:var(--pos)}
.sig-neg{background:var(--neg-bg);color:var(--neg)}
.sig-neu{background:#f4f4f4;color:var(--text3)}
</style>
</head>
<body>
<div class="hdr">
  <div class="hdr-brand">
    <div style="width:32px;height:32px;background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.18);
      border-radius:7px;display:flex;align-items:center;justify-content:center;flex-shrink:0">
      <svg width="16" height="16" viewBox="0 0 16 16" fill="none">
        <path d="M2 12 L6 6 L9 9 L13 3" stroke="white" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/>
        <circle cx="13" cy="3" r="1.5" fill="white"/>
      </svg>
    </div>
    <div>
      <h1>基金重仓回测报告</h1>
      <div class="hdr-sub">Fund Holdings Backtest</div>
    </div>
  </div>
  <span class="meta" id="meta-str"></span>
</div>

<div class="tabs">
  <button class="tb active" onclick="sw('t1')">总览</button>
  <button class="tb" onclick="sw('t2')">方向性分析</button>
  <button class="tb" onclick="sw('t3')">分位数分析</button>
  <button class="tb" onclick="sw('t4')">专项策略</button>
  <button class="tb" onclick="sw('t5')">行业 & 季度</button>
  <button class="tb" onclick="sw('t6')">明细数据</button>
</div>

<!-- ── TAB 1: 总览 ── -->
<div id="t1" class="pane active">
  <div class="cards" id="meta-cards"></div>
  <div class="g2">
    <div class="box">
      <div class="box-title">策略方向性 · T+20 超额收益对比（%）</div>
      <div id="c-overview" style="height:280px"></div>
    </div>
    <div class="box">
      <div class="box-title">fcc 分位数 · T+20 超额收益（%）</div>
      <div id="c-quintile-overview" style="height:280px"></div>
    </div>
  </div>
  <div class="tw mb">
    <table id="overview-tbl">
      <thead><tr>
        <th class="l">策略</th><th class="l">信号</th><th>N</th>
        <th>T+5 超额%</th><th>T+20 超额%</th><th>T+60 超额%</th>
        <th>胜率(T+20)</th><th>t统计量</th>
      </tr></thead>
      <tbody id="overview-body"></tbody>
    </table>
  </div>
  <div class="note">
    <b>方法论：</b>T=0 为季报法定披露截止日（保守估计，避免未来信息偏差）。
    超额收益 = 个股收益 - 同期基准（A股: 万得全A；H股: 恒生指数）。
    数据来源：基金季度持仓 × 日收盘价（前复权）。
  </div>
</div>

<!-- ── TAB 2: 方向性分析 ── -->
<div id="t2" class="pane">
  <div class="cards" id="dir-cards"></div>
  <div class="g3">
    <div class="box">
      <div class="box-title">加仓 vs 减仓 · 各持有期超额收益（%）</div>
      <div id="c-dir-bar" style="height:300px"></div>
    </div>
    <div class="box">
      <div class="box-title">胜率对比（超额收益 > 0 的比例）</div>
      <div id="c-dir-hit" style="height:300px"></div>
    </div>
    <div class="box">
      <div class="box-title">多空价差（加仓均值 - 减仓均值）</div>
      <div id="c-dir-spread" style="height:300px"></div>
    </div>
  </div>
  <div class="tw mb">
    <table>
      <thead><tr>
        <th class="l">方向</th><th>N</th>
        <th>T+5 超额%</th><th>T+20 超额%</th><th>T+60 超额%</th>
        <th>T+5 胜率</th><th>T+20 胜率</th><th>T+60 胜率</th>
        <th>T+20 t统计</th>
      </tr></thead>
      <tbody id="dir-body"></tbody>
    </table>
  </div>
  <div class="note">
    <b>加仓</b>（fcc &gt; 0）：本季度净新增持有基金数 &gt; 0 的股票。
    <b>减仓</b>（fcc &lt; 0）：净减少基金数 &lt; 0 的股票。
    <b>t统计量</b>：检验超额收益均值是否显著不为零（绝对值 &gt; 1.96 ≈ 95% 置信度）。
  </div>
</div>

<!-- ── TAB 3: 分位数分析 ── -->
<div id="t3" class="pane">
  <div class="g2">
    <div>
      <div class="box mb">
        <div class="box-title">Strategy B1 · fcc 分位数 → T+20 超额收益（%）</div>
        <div id="c-q-fcc" style="height:260px"></div>
      </div>
      <div class="box mb">
        <div class="box-title">Strategy B2 · fw_delta 分位数 → T+20 超额收益（%）</div>
        <div id="c-q-fwd" style="height:260px"></div>
      </div>
    </div>
    <div>
      <div class="box mb">
        <div class="box-title">Strategy C · 综合信号（fcc排名 + fw_delta排名）→ T+20</div>
        <div id="c-q-combo" style="height:260px"></div>
      </div>
      <div class="box mb">
        <div class="box-title">各持有期 Q5-Q1 价差（最强加仓 vs 最强减仓）</div>
        <div id="c-q-spread" style="height:260px"></div>
      </div>
    </div>
  </div>
  <div class="g2">
    <div class="tw">
      <table id="q-fcc-tbl">
        <thead><tr>
          <th class="l">fcc 分位</th><th>N</th>
          <th>T+5 超额%</th><th>T+20 超额%</th><th>T+60 超额%</th><th>胜率</th>
        </tr></thead>
        <tbody id="q-fcc-body"></tbody>
      </table>
    </div>
    <div class="tw">
      <table id="q-fwd-tbl">
        <thead><tr>
          <th class="l">fw_delta 分位</th><th>N</th>
          <th>T+5 超额%</th><th>T+20 超额%</th><th>T+60 超额%</th><th>胜率</th>
        </tr></thead>
        <tbody id="q-fwd-body"></tbody>
      </table>
    </div>
  </div>
</div>

<!-- ── TAB 4: 专项策略 ── -->
<div id="t4" class="pane">
  <div class="g2">
    <div class="box mb">
      <div class="box-title">Strategy D · 共识阈值（fcc ≥ N 只基金同步加仓）</div>
      <div id="c-consensus" style="height:280px"></div>
    </div>
    <div class="box mb">
      <div class="box-title">Strategy E · 新晋重仓 vs 拥挤减仓 vs 拥挤加仓（T+20）</div>
      <div id="c-special" style="height:280px"></div>
    </div>
  </div>
  <div class="g2">
    <div class="box mb">
      <div class="box-title">Strategy F · 加仓幅度（fw_delta 大/中/小）· T+20 超额%</div>
      <div id="c-fw-mag-add" style="height:240px"></div>
    </div>
    <div class="box mb">
      <div class="box-title">Strategy F · 减仓幅度（fw_delta 大/中/小）· T+20 超额%</div>
      <div id="c-fw-mag-red" style="height:240px"></div>
    </div>
  </div>
  <div class="tw mb">
    <table>
      <thead><tr>
        <th class="l">策略</th><th class="l">描述</th><th>N</th>
        <th>T+5 超额%</th><th>T+20 超额%</th><th>T+60 超额%</th>
        <th>胜率(T+20)</th><th>t统计量</th>
      </tr></thead>
      <tbody id="special-body"></tbody>
    </table>
  </div>
</div>

<!-- ── TAB 5: 行业 & 季度 ── -->
<div id="t5" class="pane">
  <div class="ctrl">
    <span class="lbl">持有期</span>
    <select id="ind-h" onchange="renderInd()">
      <option value="20">T+20</option>
      <option value="5">T+5</option>
      <option value="60">T+60</option>
    </select>
  </div>
  <div class="g2">
    <div class="box mb">
      <div class="box-title">各行业 加仓超额收益（%）</div>
      <div id="c-ind-add" style="height:360px"></div>
    </div>
    <div class="box mb">
      <div class="box-title">各行业 加仓-减仓 超额价差（pp）</div>
      <div id="c-ind-spread" style="height:360px"></div>
    </div>
  </div>
  <div class="g2">
    <div class="box mb">
      <div class="box-title">各季度 方向性信号有效性（加仓 T+20 超额%）</div>
      <div id="c-q-eff" style="height:260px"></div>
    </div>
    <div class="box mb">
      <div class="box-title">各季度 多空价差趋势（T+20）</div>
      <div id="c-q-trend" style="height:260px"></div>
    </div>
  </div>
  <div class="tw mb">
    <table id="ind-tbl">
      <thead><tr>
        <th class="l">行业</th><th>N(加仓)</th><th>N(减仓)</th>
        <th>加仓超额%</th><th>加仓胜率</th><th>减仓超额%</th><th>多空价差</th>
      </tr></thead>
      <tbody id="ind-body"></tbody>
    </table>
  </div>
</div>

<!-- ── TAB 6: 明细数据 ── -->
<div id="t6" class="pane">
  <div class="ctrl">
    <span class="lbl">季度</span>
    <select id="d-q"><option value="">全部</option></select>
    <span class="lbl">方向</span>
    <select id="d-dir">
      <option value="">全部</option>
      <option value="add">加仓</option>
      <option value="flat">持平</option>
      <option value="reduce">减仓</option>
    </select>
    <span class="lbl">市场</span>
    <select id="d-mkt">
      <option value="">全部</option>
      <option value="a">A股</option>
      <option value="h">H股</option>
    </select>
    <span class="lbl">搜索</span>
    <input type="text" id="d-search" placeholder="代码 / 名称" style="width:150px" oninput="renderDetail()">
    <span id="d-count" class="lbl" style="margin-left:8px"></span>
  </div>
  <div class="tw">
    <table id="detail-tbl">
      <thead><tr>
        <th class="l" onclick="dsort('q')">季度</th>
        <th class="l" onclick="dsort('code')">代码</th>
        <th class="l" onclick="dsort('name')">名称</th>
        <th class="l" onclick="dsort('ind')">行业</th>
        <th onclick="dsort('fcc')">fcc</th>
        <th onclick="dsort('fw_delta')">fw变动(pp)</th>
        <th onclick="dsort('fp')">流通股占比%</th>
        <th onclick="dsort('fc')">基金数</th>
        <th onclick="dsort('xret5')">超额T+5%</th>
        <th onclick="dsort('xret20')">超额T+20%</th>
        <th onclick="dsort('xret60')">超额T+60%</th>
      </tr></thead>
      <tbody id="detail-body"></tbody>
    </table>
  </div>
</div>

<script>
const D = __DATA__;
const H = [5, 20, 60];
const AX = {
  axisLine:{lineStyle:{color:'#ccc8bf'}},
  axisLabel:{color:'#96a0ae',fontSize:10,fontFamily:"'JetBrains Mono',monospace"},
  splitLine:{lineStyle:{color:'#e8e3db',type:'dashed'}},
};
const TT = {backgroundColor:'#fff',borderColor:'#ddd8d0',borderWidth:1,
  textStyle:{color:'#14213d',fontSize:11}};
const PAL = ['#1e50a2','#c0392b','#166534','#d97706','#7c3aed','#0891b2'];
const POS_COL = '#c0392b', NEG_COL = '#166534';

function ec(id){
  const el=document.getElementById(id); if(!el) return null;
  let c=echarts.getInstanceByDom(el); if(c) c.dispose();
  return echarts.init(el,null,{renderer:'canvas'});
}

function pct(v,d=2){
  if(v==null||isNaN(v)) return '—';
  return (v*100).toFixed(d)+'%';
}
function fmt(v,d=2){if(v==null||isNaN(v))return '—'; return v.toFixed(d);}
function sign(v){return (v!=null&&v>0)?'+':'';}
function cls(v){return v==null?'neu':v>0?'pos':v<0?'neg':'neu';}
function bar_color(v){return v>=0?POS_COL:NEG_COL;}

function sw(id){
  document.querySelectorAll('.pane').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tb').forEach(b=>b.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  const idx=['t1','t2','t3','t4','t5','t6'].indexOf(id);
  document.querySelectorAll('.tb')[idx].classList.add('active');
  const renders={t2:renderDir,t3:renderQ,t4:renderSpecial,t5:renderInd,t6:initDetail};
  if(renders[id]) renders[id]();
}

// ── helpers ─────────────────────────────────────────────────────

function statsRow(s, h, pctScale=true){
  if(!s||s.n===0) return '<td>—</td><td>—</td><td>—</td><td>—</td><td>—</td>';
  const m = s.mean; const fmt_m = pctScale ? (m*100).toFixed(2) : m.toFixed(2);
  return `<td class="${cls(m)}">${sign(m)}${fmt_m}</td>`;
}

function getStats(obj, h){ return obj && obj[`s${h}`] ? obj[`s${h}`] : null; }
function getMean(s){ return s && s.mean != null ? s.mean : null; }

function barOpt(cats, vals, color, title=''){
  const colors = Array.isArray(color) ? color : vals.map(v=>bar_color(v));
  return {
    backgroundColor:'transparent',
    grid:{left:60,right:16,top:28,bottom:36},
    xAxis:{type:'category',data:cats,...AX,axisTick:{show:false}},
    yAxis:{type:'value',...AX,axisLine:{show:false},
      axisLabel:{...AX.axisLabel,formatter:v=>(v*100).toFixed(1)+'%'}},
    tooltip:{...TT,trigger:'axis',formatter:p=>`${p[0].name}<br>${(p[0].value*100).toFixed(2)}%`},
    series:[{type:'bar',data:vals.map((v,i)=>({value:v,itemStyle:{color:colors[i],borderRadius:[3,3,0,0]}})),
      barMaxWidth:40,label:{show:true,position:'top',color:'#96a0ae',fontSize:10,
        formatter:p=>(p.value*100).toFixed(2)+'%'}}]
  };
}

function lineOpt(series, colors){
  return {
    backgroundColor:'transparent',
    grid:{left:60,right:16,top:28,bottom:36},
    xAxis:{type:'category',data:series[0].data.map((_,i)=>i),...AX},
    yAxis:{type:'value',...AX,axisLine:{show:false},
      axisLabel:{...AX.axisLabel,formatter:v=>(v*100).toFixed(1)+'%'}},
    tooltip:{...TT,trigger:'axis'},
    series:series.map((s,i)=>({name:s.name,type:'line',data:s.data,
      smooth:.4,symbolSize:6,lineStyle:{color:colors[i],width:2.5},
      itemStyle:{color:colors[i]}})),
    legend:{top:4,textStyle:{color:'#566070',fontSize:10}},
  };
}

// ── TAB 1: 总览 ─────────────────────────────────────────────────

(function initOverview(){
  const m = D.meta;
  document.getElementById('meta-str').textContent =
    `季度: ${m.quarters[0]} ~ ${m.quarters[m.quarters.length-1]}  |  生成: ${m.generated}`;

  document.getElementById('meta-cards').innerHTML = [
    ['季度数',m.quarters.length,m.quarters.join(' ')],
    ['总观测数',m.n_total.toLocaleString(),'stock-quarter pairs'],
    ['有效回测(T+20)',m.n_valid20.toLocaleString(),'有超额收益计算'],
    ['A股基准',m.bench_a,''],
    ['H股基准',m.bench_h,''],
  ].map(([l,v,s])=>`<div class="card">
    <div class="c-lbl">${l}</div>
    <div class="c-val" style="font-size:18px">${v}</div>
    <div class="c-sub">${s}</div></div>`).join('');

  // 方向性总览图
  const sa = D.strat_a;
  const cats = ['加仓组','持平组','减仓组'];
  const keys = ['add','flat','reduce'];
  const vals20 = keys.map(k=>getMean(getStats(sa[k],20)));
  const c1 = ec('c-overview');
  if(c1) c1.setOption(barOpt(cats, vals20, vals20.map(bar_color)));

  // fcc quintile 总览
  const qb = D.strat_b_fcc;
  if(qb && qb.length){
    const qcats = qb.map(q=>q.label);
    const qvals = qb.map(q=>getMean(getStats(q,20)));
    const c2 = ec('c-quintile-overview');
    if(c2) c2.setOption(barOpt(qcats,qvals,qvals.map(bar_color)));
  }

  // 总览表
  const rows = [];
  // Strategy A
  ['add','flat','reduce'].forEach((k,i)=>{
    const nm = ['加仓(fcc>0)','持平(fcc=0)','减仓(fcc<0)'][i];
    const s = sa[k];
    if(!s) return;
    const s20 = getStats(s,20);
    rows.push([`策略A·方向`,nm, s.n,
      getMean(getStats(s,5)), getMean(getStats(s,20)), getMean(getStats(s,60)),
      s20 && s20.hit, s20 && s20.tstat]);
  });
  // Strategy B (Q5 vs Q1)
  const sb = D.strat_b_fcc;
  if(sb && sb.length>=5){
    ['Q5','Q1'].forEach(lbl=>{
      const q = sb.find(x=>x.label===lbl);
      if(!q) return;
      const s20=getStats(q,20);
      rows.push([`策略B1·fcc分位`,lbl, q.n,
        getMean(getStats(q,5)), getMean(getStats(q,20)), getMean(getStats(q,60)),
        s20&&s20.hit, s20&&s20.tstat]);
    });
  }
  // Strategy D consensus fcc>=5
  const sd = D.strat_d;
  if(sd && sd['5']){
    const s20=getStats(sd['5'],20);
    rows.push([`策略D·共识≥5只`,`fcc≥5`,sd['5'].n,
      getMean(getStats(sd['5'],5)), getMean(getStats(sd['5'],20)), getMean(getStats(sd['5'],60)),
      s20&&s20.hit, s20&&s20.tstat]);
  }
  // Strategy E new entrant
  const se = D.strat_e;
  if(se && se.fresh){
    const s20=getStats(se.fresh,20);
    rows.push([`策略E·新晋`,`新晋重仓`,se.fresh.n,
      getMean(getStats(se.fresh,5)),getMean(getStats(se.fresh,20)),getMean(getStats(se.fresh,60)),
      s20&&s20.hit, s20&&s20.tstat]);
  }
  if(se && se.crowd_exit){
    const s20=getStats(se.crowd_exit,20);
    rows.push([`策略E·拥挤减仓`,`拥挤出逃`,se.crowd_exit.n,
      getMean(getStats(se.crowd_exit,5)),getMean(getStats(se.crowd_exit,20)),getMean(getStats(se.crowd_exit,60)),
      s20&&s20.hit,s20&&s20.tstat]);
  }

  document.getElementById('overview-body').innerHTML = rows.map(
    ([strat,sig,n,r5,r20,r60,hit,t])=>{
      const tc = t==null?'neu':Math.abs(t)>=1.96?'pos':'neu';
      return `<tr>
        <td class="l">${strat}</td>
        <td class="l"><span class="tag">${sig}</span></td>
        <td>${n||'—'}</td>
        <td class="${cls(r5)}">${r5!=null?sign(r5)+(r5*100).toFixed(2)+'%':'—'}</td>
        <td class="${cls(r20)}">${r20!=null?sign(r20)+(r20*100).toFixed(2)+'%':'—'}</td>
        <td class="${cls(r60)}">${r60!=null?sign(r60)+(r60*100).toFixed(2)+'%':'—'}</td>
        <td>${hit!=null?(hit*100).toFixed(1)+'%':'—'}</td>
        <td class="${tc}">${t!=null?t.toFixed(2):'—'}</td>
      </tr>`;
    }).join('');
})();

// ── TAB 2: 方向性 ───────────────────────────────────────────────

function renderDir(){
  const sa = D.strat_a;
  const cats = ['T+5','T+20','T+60'];
  const add_vals = H.map(h=>getMean(getStats(sa.add,h)));
  const red_vals = H.map(h=>getMean(getStats(sa.reduce,h)));
  const flat_vals = H.map(h=>getMean(getStats(sa.flat,h)));

  // Cards
  const s20a = getStats(sa.add,20)||{};
  const s20r = getStats(sa.reduce,20)||{};
  const spread = (s20a.mean!=null&&s20r.mean!=null) ? s20a.mean-s20r.mean : null;
  document.getElementById('dir-cards').innerHTML = [
    ['加仓组 N', (sa.add||{}).n||'—', 'fcc > 0'],
    ['减仓组 N', (sa.reduce||{}).n||'—', 'fcc < 0'],
    ['加仓 T+20', s20a.mean!=null?(sign(s20a.mean)+(s20a.mean*100).toFixed(2)+'%'):'—','超额收益均值'],
    ['减仓 T+20', s20r.mean!=null?(sign(s20r.mean)+(s20r.mean*100).toFixed(2)+'%'):'—','超额收益均值'],
    ['多空价差', spread!=null?((spread*100).toFixed(2)+'pp'):'—','加仓-减仓 T+20'],
  ].map(([l,v,s])=>`<div class="card">
    <div class="c-lbl">${l}</div><div class="c-val" style="font-size:16px">${v}</div>
    <div class="c-sub">${s}</div></div>`).join('');

  // Bar: returns by direction
  const c1=ec('c-dir-bar');
  if(c1) c1.setOption({
    backgroundColor:'transparent',grid:{left:60,right:16,top:36,bottom:36},
    xAxis:{type:'category',data:cats,...AX,axisTick:{show:false}},
    yAxis:{type:'value',...AX,axisLine:{show:false},
      axisLabel:{...AX.axisLabel,formatter:v=>(v*100).toFixed(1)+'%'}},
    tooltip:{...TT,trigger:'axis'},
    legend:{top:4,textStyle:{color:'#566070',fontSize:10}},
    series:[
      {name:'加仓',type:'bar',data:add_vals,barMaxWidth:28,itemStyle:{color:POS_COL,borderRadius:[3,3,0,0]}},
      {name:'持平',type:'bar',data:flat_vals,barMaxWidth:28,itemStyle:{color:'#aab',borderRadius:[3,3,0,0]}},
      {name:'减仓',type:'bar',data:red_vals,barMaxWidth:28,itemStyle:{color:NEG_COL,borderRadius:[3,3,0,0]}},
    ]
  });

  // Hit rate
  const add_hit = H.map(h=>{ const s=getStats(sa.add,h); return s?s.hit:null; });
  const red_hit = H.map(h=>{ const s=getStats(sa.reduce,h); return s?s.hit:null; });
  const c2=ec('c-dir-hit');
  if(c2) c2.setOption({
    backgroundColor:'transparent',grid:{left:60,right:16,top:36,bottom:36},
    xAxis:{type:'category',data:cats,...AX,axisTick:{show:false}},
    yAxis:{type:'value',min:0.3,max:0.7,...AX,axisLine:{show:false},
      axisLabel:{...AX.axisLabel,formatter:v=>(v*100).toFixed(0)+'%'}},
    tooltip:{...TT,trigger:'axis',formatter:p=>p.map(s=>`${s.seriesName}: ${(s.value*100).toFixed(1)}%`).join('<br>')},
    legend:{top:4,textStyle:{color:'#566070',fontSize:10}},
    series:[
      {name:'加仓胜率',type:'line',data:add_hit,smooth:.4,symbolSize:6,
        lineStyle:{color:POS_COL,width:2.5},itemStyle:{color:POS_COL}},
      {name:'减仓胜率',type:'line',data:red_hit,smooth:.4,symbolSize:6,
        lineStyle:{color:NEG_COL,width:2.5},itemStyle:{color:NEG_COL}},
    ]
  });

  // Spread
  const spread_vals = H.map(h=>{
    const a=getMean(getStats(sa.add,h)); const r=getMean(getStats(sa.reduce,h));
    return (a!=null&&r!=null)?a-r:null;
  });
  const c3=ec('c-dir-spread');
  if(c3) c3.setOption(barOpt(['T+5','T+20','T+60'],spread_vals,spread_vals.map(bar_color)));

  // Table
  document.getElementById('dir-body').innerHTML = [
    ['加仓 (fcc>0)','add'],['持平 (fcc=0)','flat'],['减仓 (fcc<0)','reduce'],
  ].map(([lbl,k])=>{
    const grp=sa[k]||{};
    const s5=getStats(grp,5)||{}, s20=getStats(grp,20)||{}, s60=getStats(grp,60)||{};
    return `<tr>
      <td class="l"><b>${lbl}</b></td><td>${grp.n||0}</td>
      <td class="${cls(s5.mean)}">${s5.mean!=null?sign(s5.mean)+(s5.mean*100).toFixed(2)+'%':'—'}</td>
      <td class="${cls(s20.mean)}">${s20.mean!=null?sign(s20.mean)+(s20.mean*100).toFixed(2)+'%':'—'}</td>
      <td class="${cls(s60.mean)}">${s60.mean!=null?sign(s60.mean)+(s60.mean*100).toFixed(2)+'%':'—'}</td>
      <td>${s5.hit!=null?(s5.hit*100).toFixed(1)+'%':'—'}</td>
      <td>${s20.hit!=null?(s20.hit*100).toFixed(1)+'%':'—'}</td>
      <td>${s60.hit!=null?(s60.hit*100).toFixed(1)+'%':'—'}</td>
      <td class="${s20.tstat&&Math.abs(s20.tstat)>=1.96?'pos':'neu'}">${s20.tstat!=null?s20.tstat.toFixed(2):'—'}</td>
    </tr>`;
  }).join('');
}

// ── TAB 3: 分位数 ────────────────────────────────────────────────

function renderQ(){
  function quintileBar(containerId, data, label){
    if(!data||!data.length) return;
    const c=ec(containerId); if(!c) return;
    const cats=data.map(d=>d.label);
    const v5=data.map(d=>getMean(getStats(d,5)));
    const v20=data.map(d=>getMean(getStats(d,20)));
    const v60=data.map(d=>getMean(getStats(d,60)));
    c.setOption({
      backgroundColor:'transparent',
      grid:{left:60,right:16,top:40,bottom:36},
      xAxis:{type:'category',data:cats,...AX,axisTick:{show:false}},
      yAxis:{type:'value',...AX,axisLine:{show:false},
        axisLabel:{...AX.axisLabel,formatter:v=>(v*100).toFixed(1)+'%'}},
      tooltip:{...TT,trigger:'axis',formatter:p=>p.map(s=>`${s.seriesName}: ${(s.value*100).toFixed(2)}%`).join('<br>')},
      legend:{top:4,textStyle:{color:'#566070',fontSize:10}},
      series:[
        {name:'T+5',type:'bar',data:v5.map((v,i)=>({value:v,itemStyle:{color:v>=0?POS_COL:NEG_COL,opacity:.55,borderRadius:[2,2,0,0]}})),barMaxWidth:20},
        {name:'T+20',type:'bar',data:v20.map((v,i)=>({value:v,itemStyle:{color:v>=0?POS_COL:NEG_COL,borderRadius:[2,2,0,0]}})),barMaxWidth:20},
        {name:'T+60',type:'bar',data:v60.map((v,i)=>({value:v,itemStyle:{color:v>=0?POS_COL:NEG_COL,opacity:.7,borderRadius:[2,2,0,0]}})),barMaxWidth:20},
      ]
    });
  }

  quintileBar('c-q-fcc', D.strat_b_fcc, 'fcc');
  quintileBar('c-q-fwd', D.strat_b_fwd, 'fw_delta');
  quintileBar('c-q-combo', D.strat_c, '综合');

  // Q5-Q1 spread across horizons
  function spread(data){
    if(!data||data.length<5) return H.map(()=>null);
    const q5=data.find(d=>d.label==='Q5'||d.label==='C5');
    const q1=data.find(d=>d.label==='Q1'||d.label==='C1');
    return H.map(h=>{
      const a=getMean(getStats(q5,h)), b=getMean(getStats(q1,h));
      return a!=null&&b!=null?a-b:null;
    });
  }
  const c=ec('c-q-spread'); if(c){
    const sf=spread(D.strat_b_fcc), sw=spread(D.strat_b_fwd), sc=spread(D.strat_c);
    c.setOption({
      backgroundColor:'transparent',grid:{left:60,right:16,top:40,bottom:36},
      xAxis:{type:'category',data:['T+5','T+20','T+60'],...AX,axisTick:{show:false}},
      yAxis:{type:'value',...AX,axisLine:{show:false},
        axisLabel:{...AX.axisLabel,formatter:v=>(v*100).toFixed(1)+'%'}},
      tooltip:{...TT,trigger:'axis',formatter:p=>p.map(s=>`${s.seriesName}: ${(s.value*100).toFixed(2)}pp`).join('<br>')},
      legend:{top:4,textStyle:{color:'#566070',fontSize:10}},
      series:[
        {name:'fcc Q5-Q1',type:'line',data:sf,smooth:.4,symbolSize:6,lineStyle:{color:PAL[0],width:2},itemStyle:{color:PAL[0]}},
        {name:'fw_delta Q5-Q1',type:'line',data:sw,smooth:.4,symbolSize:6,lineStyle:{color:PAL[1],width:2},itemStyle:{color:PAL[1]}},
        {name:'综合 C5-C1',type:'line',data:sc,smooth:.4,symbolSize:6,lineStyle:{color:PAL[2],width:2},itemStyle:{color:PAL[2]}},
      ]
    });
  }

  // Tables
  function fillQTable(bodyId, data){
    if(!data) return;
    document.getElementById(bodyId).innerHTML = data.map(d=>{
      const s5=getStats(d,5)||{}, s20=getStats(d,20)||{}, s60=getStats(d,60)||{};
      return `<tr>
        <td class="l"><b>${d.label}</b></td><td>${d.n}</td>
        <td class="${cls(s5.mean)}">${s5.mean!=null?sign(s5.mean)+(s5.mean*100).toFixed(2)+'%':'—'}</td>
        <td class="${cls(s20.mean)}">${s20.mean!=null?sign(s20.mean)+(s20.mean*100).toFixed(2)+'%':'—'}</td>
        <td class="${cls(s60.mean)}">${s60.mean!=null?sign(s60.mean)+(s60.mean*100).toFixed(2)+'%':'—'}</td>
        <td>${s20.hit!=null?(s20.hit*100).toFixed(1)+'%':'—'}</td>
      </tr>`;
    }).join('');
  }
  fillQTable('q-fcc-body', D.strat_b_fcc);
  fillQTable('q-fwd-body', D.strat_b_fwd);
}

// ── TAB 4: 专项策略 ─────────────────────────────────────────────

function renderSpecial(){
  // Consensus chart
  const sd = D.strat_d;
  if(sd){
    const thrs = Object.keys(sd).map(Number).sort((a,b)=>a-b);
    const c=ec('c-consensus'); if(c){
      c.setOption({
        backgroundColor:'transparent',grid:{left:60,right:16,top:40,bottom:36},
        xAxis:{type:'category',data:thrs.map(t=>`≥${t}只`), ...AX,axisTick:{show:false}},
        yAxis:{type:'value',...AX,axisLine:{show:false},
          axisLabel:{...AX.axisLabel,formatter:v=>(v*100).toFixed(1)+'%'}},
        tooltip:{...TT,trigger:'axis',formatter:p=>p.map(s=>`${s.seriesName}: ${(s.value*100).toFixed(2)}%`).join('<br>')},
        legend:{top:4,textStyle:{color:'#566070',fontSize:10}},
        series:H.map((h,i)=>({
          name:`T+${h}`,type:'line',
          data:thrs.map(t=>getMean(getStats(sd[String(t)],h))),
          smooth:.4,symbolSize:6,lineStyle:{color:PAL[i],width:2},itemStyle:{color:PAL[i]},
        }))
      });
    }
  }

  // Special strategies comparison
  const se = D.strat_e||{};
  const specials = [
    {key:'fresh',label:'新晋重仓 (fcc≥5 & fp<2%)',color:PAL[0]},
    {key:'crowd_add',label:'拥挤加仓 (fp>3% & fcc≥3)',color:PAL[3]},
    {key:'crowd_exit',label:'拥挤减仓 (fp>3% & fcc≤-3)',color:PAL[1]},
  ];
  const c2=ec('c-special'); if(c2){
    c2.setOption({
      backgroundColor:'transparent',grid:{left:56,right:16,top:40,bottom:80},
      xAxis:{type:'category',data:specials.map(s=>s.label),...AX,axisTick:{show:false},
        axisLabel:{...AX.axisLabel,rotate:10,width:120,overflow:'break'}},
      yAxis:{type:'value',...AX,axisLine:{show:false},
        axisLabel:{...AX.axisLabel,formatter:v=>(v*100).toFixed(1)+'%'}},
      tooltip:{...TT,trigger:'axis',formatter:p=>p.map(s=>`${s.seriesName}: ${(s.value*100).toFixed(2)}%`).join('<br>')},
      legend:{top:4,textStyle:{color:'#566070',fontSize:10}},
      series:H.map((h,i)=>({
        name:`T+${h}`,type:'bar',barMaxWidth:24,
        data:specials.map(s=>getMean(getStats(se[s.key],h))),
        itemStyle:{color:PAL[i],borderRadius:[3,3,0,0]},
      }))
    });
  }

  // fw magnitude charts
  const sf = D.strat_f||{};
  function magChart(id, data){
    if(!data||!Object.keys(data).length) return;
    const c=ec(id); if(!c) return;
    const cats = Object.keys(data);
    c.setOption(barOpt(cats, cats.map(k=>getMean(getStats(data[k],20))),
      cats.map(k=>bar_color(getMean(getStats(data[k],20))))));
  }
  magChart('c-fw-mag-add', sf.add);
  magChart('c-fw-mag-red', sf.reduce);

  // Summary table
  const rows = [];
  // Consensus
  if(sd){
    [1,3,5,10].forEach(t=>{
      const grp=sd[String(t)]; if(!grp) return;
      const s20=getStats(grp,20)||{};
      rows.push(['策略D·共识',`fcc ≥ ${t} 只`,grp.n,
        getMean(getStats(grp,5)),getMean(getStats(grp,20)),getMean(getStats(grp,60)),
        s20.hit,s20.tstat]);
    });
  }
  // Special
  const specialLabels={fresh:'新晋重仓', crowd_add:'拥挤加仓', crowd_exit:'拥挤减仓'};
  const specialDescs={fresh:'fcc≥5 & fp<2%', crowd_add:'fp>3% & fcc≥3', crowd_exit:'fp>3% & fcc≤-3'};
  Object.keys(specialLabels).forEach(k=>{
    const grp=se[k]; if(!grp) return;
    const s20=getStats(grp,20)||{};
    rows.push(['策略E·专项',specialLabels[k]+'('+specialDescs[k]+')',grp.n,
      getMean(getStats(grp,5)),getMean(getStats(grp,20)),getMean(getStats(grp,60)),
      s20.hit,s20.tstat]);
  });
  // fw magnitude
  ['add','reduce'].forEach(dir=>{
    const data=(sf||{})[dir]||{};
    Object.keys(data).forEach(sz=>{
      const grp=data[sz]; const s20=getStats(grp,20)||{};
      rows.push([`策略F·${dir==='add'?'加':'减'}仓幅度`,
        `${sz}幅${dir==='add'?'加':'减'}仓`,grp.n,
        getMean(getStats(grp,5)),getMean(getStats(grp,20)),getMean(getStats(grp,60)),
        s20.hit,s20.tstat]);
    });
  });

  document.getElementById('special-body').innerHTML = rows.map(
    ([strat,desc,n,r5,r20,r60,hit,t])=>`<tr>
      <td class="l">${strat}</td><td class="l" style="font-size:11px;color:#566070">${desc}</td>
      <td>${n||'—'}</td>
      <td class="${cls(r5)}">${r5!=null?sign(r5)+(r5*100).toFixed(2)+'%':'—'}</td>
      <td class="${cls(r20)}">${r20!=null?sign(r20)+(r20*100).toFixed(2)+'%':'—'}</td>
      <td class="${cls(r60)}">${r60!=null?sign(r60)+(r60*100).toFixed(2)+'%':'—'}</td>
      <td>${hit!=null?(hit*100).toFixed(1)+'%':'—'}</td>
      <td class="${t&&Math.abs(t)>=1.96?'pos':'neu'}">${t!=null?t.toFixed(2):'—'}</td>
    </tr>`).join('');
}

// ── TAB 5: 行业 & 季度 ──────────────────────────────────────────

function renderInd(){
  const h = parseInt(document.getElementById('ind-h').value)||20;
  const bi = D.by_ind||{};
  const bq = D.by_q||{};

  // Industry charts
  let inds = Object.entries(bi).filter(([,v])=>v.n_add>=3||v.n_red>=3);
  inds.sort((a,b)=>{
    const am=getMean(getStats(a[1],h,'add')); const bm=getMean(getStats(b[1],h,'add'));
    return (bm||0)-(am||0);
  });
  const maxInds = Math.min(inds.length, 20);

  function indMean(v,h,dir){
    const s = v[`${dir}_s${h}`];
    return s&&s.mean!=null?s.mean:null;
  }

  // Industry add chart
  const c1=ec('c-ind-add'); if(c1){
    const slice=inds.slice(0,maxInds);
    const cats=slice.map(([ind])=>ind);
    const vals=slice.map(([,v])=>indMean(v,h,'add'));
    c1.style&&(c1.getDom().style.height=Math.max(300,cats.length*20)+'px');
    c1.setOption({
      backgroundColor:'transparent',
      grid:{left:90,right:60,top:10,bottom:30},
      xAxis:{type:'value',...AX,axisLine:{show:false},
        axisLabel:{...AX.axisLabel,formatter:v=>(v*100).toFixed(1)+'%'}},
      yAxis:{type:'category',data:cats,inverse:false,...AX,axisLine:{show:false},axisTick:{show:false},
        axisLabel:{color:'#566070',fontSize:10,width:80,overflow:'truncate',fontFamily:"'Sora',sans-serif"}},
      tooltip:{...TT,trigger:'axis',formatter:p=>`${p[0].name}: ${(p[0].value*100).toFixed(2)}%`},
      series:[{type:'bar',data:vals.map(v=>({value:v,itemStyle:{color:bar_color(v),borderRadius:[0,3,3,0]}})),
        barMaxWidth:18,
        label:{show:true,position:'right',color:'#96a0ae',fontSize:9,
          formatter:p=>(p.value*100).toFixed(2)+'%'}}]
    });
  }

  // Industry spread chart
  const c2=ec('c-ind-spread'); if(c2){
    const slice=inds.slice(0,maxInds);
    const cats=slice.map(([ind])=>ind);
    const spreads=slice.map(([,v])=>{
      const a=indMean(v,h,'add'), r=indMean(v,h,'red');
      return a!=null&&r!=null?a-r:null;
    });
    c2.setOption({
      backgroundColor:'transparent',
      grid:{left:90,right:60,top:10,bottom:30},
      xAxis:{type:'value',...AX,axisLine:{show:false},
        axisLabel:{...AX.axisLabel,formatter:v=>(v*100).toFixed(1)+'%'}},
      yAxis:{type:'category',data:cats,...AX,axisLine:{show:false},axisTick:{show:false},
        axisLabel:{color:'#566070',fontSize:10,width:80,overflow:'truncate',fontFamily:"'Sora',sans-serif"}},
      tooltip:{...TT,trigger:'axis',formatter:p=>`${p[0].name}: ${(p[0].value*100).toFixed(2)}pp`},
      series:[{type:'bar',data:spreads.map(v=>({value:v,itemStyle:{color:bar_color(v),borderRadius:[0,3,3,0]}})),
        barMaxWidth:18,
        label:{show:true,position:'right',color:'#96a0ae',fontSize:9,
          formatter:p=>(p.value*100).toFixed(2)+'pp'}}]
    });
  }

  // Quarter effectiveness
  const qs = D.meta.quarters;
  const c3=ec('c-q-eff'); if(c3){
    const addVals=qs.map(q=>{const bq_=bq[q]; return bq_?getMean(getStats(bq_,20,'add')):null;});
    const redVals=qs.map(q=>{const bq_=bq[q]; return bq_?getMean(getStats(bq_,20,'red')):null;});
    c3.setOption({
      backgroundColor:'transparent',grid:{left:60,right:16,top:40,bottom:36},
      xAxis:{type:'category',data:qs,...AX,axisTick:{show:false}},
      yAxis:{type:'value',...AX,axisLine:{show:false},
        axisLabel:{...AX.axisLabel,formatter:v=>(v*100).toFixed(1)+'%'}},
      tooltip:{...TT,trigger:'axis',formatter:p=>p.map(s=>`${s.seriesName}: ${(s.value*100).toFixed(2)}%`).join('<br>')},
      legend:{top:4,textStyle:{color:'#566070',fontSize:10}},
      series:[
        {name:'加仓超额',type:'bar',data:addVals.map(v=>({value:v,itemStyle:{color:POS_COL,borderRadius:[3,3,0,0]}})),barMaxWidth:24},
        {name:'减仓超额',type:'bar',data:redVals.map(v=>({value:v,itemStyle:{color:NEG_COL,borderRadius:[3,3,0,0]}})),barMaxWidth:24},
      ]
    });
  }

  // Quarter trend (L-S spread)
  const c4=ec('c-q-trend'); if(c4){
    const spreads=qs.map(q=>{
      const bq_=bq[q]; if(!bq_) return null;
      const a=getMean(getStats(bq_,20,'add')), r=getMean(getStats(bq_,20,'red'));
      return a!=null&&r!=null?a-r:null;
    });
    c4.setOption({
      backgroundColor:'transparent',grid:{left:60,right:16,top:28,bottom:36},
      xAxis:{type:'category',data:qs,...AX,axisTick:{show:false}},
      yAxis:{type:'value',...AX,axisLine:{show:false},
        axisLabel:{...AX.axisLabel,formatter:v=>(v*100).toFixed(1)+'%'}},
      tooltip:{...TT,trigger:'axis',formatter:p=>`多空价差: ${(p[0].value*100).toFixed(2)}pp`},
      series:[{type:'line',data:spreads,smooth:.4,symbolSize:7,
        areaStyle:{color:{type:'linear',x:0,y:0,x2:0,y2:1,
          colorStops:[{offset:0,color:'#1e50a228'},{offset:1,color:'#1e50a204'}]}},
        lineStyle:{color:PAL[0],width:2.5},itemStyle:{color:PAL[0]}}]
    });
  }

  // Industry table
  const sortedInds = [...inds].sort((a,b)=>{
    const as=indMean(a[1],h,'add'), bs=indMean(b[1],h,'add');
    return (bs||0)-(as||0);
  });
  document.getElementById('ind-body').innerHTML = sortedInds.map(([ind,v])=>{
    const am=indMean(v,h,'add'), rm=indMean(v,h,'red');
    const sp=am!=null&&rm!=null?am-rm:null;
    const ah=v[`add_s${h}`]&&v[`add_s${h}`].hit;
    return `<tr>
      <td class="l"><span class="tag">${ind}</span></td>
      <td>${v.n_add||0}</td><td>${v.n_red||0}</td>
      <td class="${cls(am)}">${am!=null?sign(am)+(am*100).toFixed(2)+'%':'—'}</td>
      <td>${ah!=null?(ah*100).toFixed(1)+'%':'—'}</td>
      <td class="${cls(rm)}">${rm!=null?sign(rm)+(rm*100).toFixed(2)+'%':'—'}</td>
      <td class="${cls(sp)}">${sp!=null?sign(sp)+(sp*100).toFixed(2)+'pp':'—'}</td>
    </tr>`;
  }).join('');
}

// Helper: industry/quarter stats use different key pattern
(function patchStats(){
  const origGet = getStats;
  window.getStats = function(obj, h, dir){
    if(!obj) return null;
    if(dir){
      // for industry/quarter obj: obj.add_s20 or obj.red_s20
      const key = `${dir}_s${h}`;
      return obj[key]||null;
    }
    // standard: obj.s20
    return obj[`s${h}`]||null;
  };
})();

// ── TAB 6: 明细 ─────────────────────────────────────────────────

let detailSortKey='xret20', detailSortAsc=false;
let detailInited=false;

function initDetail(){
  if(!detailInited){
    const qsel=document.getElementById('d-q');
    D.meta.quarters.forEach(q=>{
      const o=document.createElement('option'); o.value=q; o.text=q; qsel.appendChild(o);
    });
    ['d-q','d-dir','d-mkt'].forEach(id=>{
      document.getElementById(id).addEventListener('change',renderDetail);
    });
    detailInited=true;
  }
  renderDetail();
}

function dsort(key){
  if(detailSortKey===key) detailSortAsc=!detailSortAsc;
  else {detailSortKey=key; detailSortAsc=false;}
  renderDetail();
}

function renderDetail(){
  const qf=document.getElementById('d-q').value;
  const df=document.getElementById('d-dir').value;
  const mf=document.getElementById('d-mkt').value;
  const sq=document.getElementById('d-search').value.trim().toLowerCase();

  let rows=D.records.filter(r=>{
    if(qf && r.q!==qf) return false;
    if(df){
      const fcc=r.fcc;
      if(df==='add' && !(fcc>0)) return false;
      if(df==='flat' && !(fcc===0)) return false;
      if(df==='reduce' && !(fcc<0)) return false;
    }
    if(mf==='a' && r.is_hk) return false;
    if(mf==='h' && !r.is_hk) return false;
    if(sq && !r.code.toLowerCase().includes(sq) && !r.name.includes(sq)) return false;
    return true;
  });

  // sort
  rows.sort((a,b)=>{
    const av=a[detailSortKey], bv=b[detailSortKey];
    if(av==null&&bv==null) return 0;
    if(av==null) return 1; if(bv==null) return -1;
    if(typeof av==='string') return detailSortAsc?av.localeCompare(bv):bv.localeCompare(av);
    return detailSortAsc?av-bv:bv-av;
  });

  document.getElementById('d-count').textContent=`共 ${rows.length} 条`;

  // Update sort indicators
  document.querySelectorAll('#detail-tbl th').forEach(th=>{
    th.classList.remove('asc','desc');
    const fn=th.getAttribute('onclick');
    if(fn&&fn.includes(`'${detailSortKey}'`)) th.classList.add(detailSortAsc?'asc':'desc');
  });

  const MAX=300;
  const slice=rows.slice(0,MAX);
  document.getElementById('detail-body').innerHTML=slice.map(r=>`<tr>
    <td class="l">${r.q}</td>
    <td class="l code">${r.code}</td>
    <td class="l">${r.name}</td>
    <td class="l"><span class="tag" style="font-size:10px">${r.ind||'—'}</span></td>
    <td class="${r.fcc>0?'pos':r.fcc<0?'neg':'neu'}">${r.fcc!=null?(r.fcc>0?'+':'')+r.fcc:'—'}</td>
    <td class="${cls(r.fw_delta)}">${r.fw_delta!=null?sign(r.fw_delta)+r.fw_delta.toFixed(3)+'pp':'—'}</td>
    <td>${r.fp!=null?(r.fp*100).toFixed(2)+'%':'—'}</td>
    <td>${r.fc!=null?r.fc:'—'}</td>
    <td class="${cls(r.xret5)}">${r.xret5!=null?sign(r.xret5)+(r.xret5*100).toFixed(2)+'%':'—'}</td>
    <td class="${cls(r.xret20)}">${r.xret20!=null?sign(r.xret20)+(r.xret20*100).toFixed(2)+'%':'—'}</td>
    <td class="${cls(r.xret60)}">${r.xret60!=null?sign(r.xret60)+(r.xret60*100).toFixed(2)+'%':'—'}</td>
  </tr>`).join('');
  if(rows.length>MAX){
    document.getElementById('detail-body').innerHTML+=
      `<tr><td colspan="11" style="text-align:center;color:#96a0ae;padding:12px">
        仅显示前 ${MAX} 条，请使用筛选缩小范围（共 ${rows.length} 条）</td></tr>`;
  }
}
</script>
</body>
</html>
"""

# ═══════════════════════════════════════════════════════════════
# 7. main
# ═══════════════════════════════════════════════════════════════

def build_html(data):
    data_json = json.dumps(data, ensure_ascii=False, separators=(",", ":"),
                           default=lambda x: None if (isinstance(x, float) and x != x) else x)
    return HTML_TEMPLATE.replace("__DATA__", data_json)


def main():
    print("=" * 60)
    print("基金重仓回测系统")
    print("=" * 60)

    quarters, fund_df = load_fund_holdings()

    print("加载价格数据...")
    a_prices  = load_price_csv(A_CSV)
    h_prices  = load_price_csv(H_CSV)
    idx_prices = load_price_csv(IDX_CSV)
    print(f"  A股: {a_prices.shape}  H股: {h_prices.shape}  指数: {idx_prices.shape}")

    df = build_dataset(quarters, fund_df, a_prices, h_prices, idx_prices)

    analysis = build_analysis(quarters, df)

    print("生成 HTML...")
    html = build_html(analysis)
    OUTPUT.write_text(html, encoding="utf-8")
    kb = OUTPUT.stat().st_size / 1024
    print(f"  → {OUTPUT}  ({kb:.0f} KB)")
    print("完成！用浏览器打开 backtest_report.html 查看报告。")
    print("=" * 60)

    # 打印关键结果摘要
    sa = analysis["strat_a"]
    print("\n── 关键结论（T+20 超额收益）──")
    for k, lbl in [("add","加仓"), ("reduce","减仓"), ("flat","持平")]:
        s = (sa.get(k) or {}).get("s20") or {}
        if s.get("mean") is not None:
            print(f"  {lbl}: {s['mean']*100:+.2f}%  胜率={s['hit']*100:.1f}%  t={s.get('tstat') or 0:.2f}  N={s['n']}")


if __name__ == "__main__":
    main()
