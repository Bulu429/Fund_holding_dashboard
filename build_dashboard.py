#!/usr/bin/env python3
"""基金重仓看板生成脚本 - 运行后刷新 fund_dashboard.html"""

import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

BASE = Path("workwork/research/策略研究/基金重仓看板")
EXCEL = BASE / "20260122 基金持仓个股视角.xlsx"
OUTPUT = BASE / "fund_dashboard.html"

# 列索引 (0-based)，基于 Row2 标题确认
COL = {
    'code':       0,   # A: 代码
    'name':       1,   # B: 名称
    'end_price':  2,   # C: 季度末股价
    'shares':     3,   # D: 持仓数量(万股)
    'shares_chg': 4,   # E: 持仓变动(万股)
    'pos_value':  5,   # F: 持仓市值(亿元) = V/10000
    'fund_weight':6,   # G: 占基金持仓比重 (decimal)
    'float_pct':  7,   # H: 占流通股比 (decimal)
    'chg_pct':    8,   # I: 变动比例 (decimal)
    'val_chg':    9,   # J: 变动市值(亿元)
    'industry':  11,   # L: 行业
    'cur_price': 13,   # N: 现价
    'price_chg': 14,   # O: 涨跌
    'price_pct': 15,   # P: 涨跌幅 (decimal)
    'r60':       16,   # Q: 60日涨跌幅 (decimal)
    'ytd':       17,   # R: 年初至今 (decimal)
    'float_cap': 18,   # S: 流通市值(元)
    'fund_count':23,   # X: 持有基金数
    'fund_chg':  24,   # Y: 基金增减数量
}

QUARTER_RE = re.compile(r'^([1-4]Q\d{2})\s*个股\s*$')

# 旧格式列映射（3Q23等早期 sheet，Row1为标题行，float_pct已是百分比形式）
COL_OLD = {
    'name': 0, 'shares': 1, 'float_pct': 2, 'shares_chg': 3,
    'chg_pct': 4, 'pos_value': 5, 'code': 6, 'industry': 7,
    'fund_count': 8, 'fund_chg': 9,
}


def pv(v):
    if v is None:
        return None
    try:
        f = float(v)
        return None if f != f else round(f, 4)
    except (TypeError, ValueError):
        return None


def sort_quarter_key(q):
    m = re.match(r'(\d)Q(\d{2})', q)
    return (int(m.group(2)), int(m.group(1))) if m else (0, 0)


def _is_old_format(ws):
    """旧格式：A1 = '名称'（字符串），无配置行"""
    first = next(ws.iter_rows(min_row=1, max_row=1, max_col=1, values_only=True), [None])[0]
    return isinstance(first, str) and '名称' in str(first)


def load_data():
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    raw = []
    quarters = set()

    for sname in wb.sheetnames:
        m = QUARTER_RE.match(sname)
        if not m:
            continue
        q = m.group(1)
        if q in ('3Q23', '4Q23'):   # 数据未处理，跳过
            continue
        quarters.add(q)
        ws = wb[sname]

        if _is_old_format(ws):
            # 旧格式：Row1=标题，Row2+=数据，float_pct 已是 %，需÷100
            for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
                code = row[COL_OLD['code']] if len(row) > COL_OLD['code'] else None
                if not code or not isinstance(code, str) or '.' not in str(code):
                    continue
                def go(key):
                    idx = COL_OLD[key]
                    return row[idx] if len(row) > idx else None
                fp_raw = pv(go('float_pct'))
                raw.append({
                    'q': q, 'code': str(code).strip(),
                    'name': str(go('name') or '').strip(),
                    'ind':  str(go('industry') or '其他').strip() or '其他',
                    'ep': None, 'sh': pv(go('shares')), 'sch': pv(go('shares_chg')),
                    'pv': pv(go('pos_value')),  # 亿元
                    'fw': None, 'vc': None,
                    'fp': round(fp_raw / 100, 6) if fp_raw is not None else None,
                    'cp': None, 'pp': None, 'r60': None, 'ytd': None,
                    'fc': pv(go('fund_count')), 'fcc': pv(go('fund_chg')),
                })
        else:
            # 新格式：Row1=配置，Row2=标题，Row3+=数据
            # 动态检测可能随 sheet 变化的列位置（如 1Q24 中 fund_count/fund_chg 列偏移不同）
            col_map = dict(COL)
            hdr = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
            if hdr:
                for i, cell in enumerate(hdr):
                    if cell == '持有基金数':
                        col_map['fund_count'] = i
                    elif cell == '基金增减数量':
                        col_map['fund_chg'] = i
            max_col = max(col_map.values()) + 1
            for row in ws.iter_rows(min_row=3, max_col=max_col, values_only=True):
                code = row[col_map['code']] if len(row) > col_map['code'] else None
                if not code or not isinstance(code, str) or '.' not in str(code):
                    continue
                def g(key):
                    idx = col_map[key]
                    return row[idx] if len(row) > idx else None
                raw.append({
                    'q': q, 'code': str(code).strip(),
                    'name': str(g('name') or '').strip(),
                    'ind':  str(g('industry') or '其他').strip() or '其他',
                    'ep':  pv(g('end_price')), 'sh': pv(g('shares')),
                    'sch': pv(g('shares_chg')), 'pv': pv(g('pos_value')),
                    'fw':  pv(g('fund_weight')), 'fp': pv(g('float_pct')),
                    'cp':  pv(g('cur_price')), 'pp': pv(g('price_pct')),
                    'r60': pv(g('r60')), 'ytd': pv(g('ytd')),
                    'fc':  pv(g('fund_count')), 'fcc': pv(g('fund_chg')),
                    'vc':  pv(g('val_chg')),    # J: 变动市值(亿元)
                })

    wb.close()
    quarters = sorted(quarters, key=sort_quarter_key)
    return quarters, raw


def build_dataset(quarters, records):
    """Pre-compute derived datasets for JS rendering."""
    by_quarter = defaultdict(list)
    for r in records:
        by_quarter[r['q']].append(r)

    # per-stock history for tab3
    by_stock = {}
    for r in records:
        code = r['code']
        if code not in by_stock:
            by_stock[code] = {'name': r['name'], 'ind': r['ind'], 'hist': []}
        by_stock[code]['hist'].append({
            'q': r['q'], 'fc': r['fc'], 'pv': r['pv'],
            'fp': r['fp'], 'fw': r['fw'], 'fcc': r['fcc'], 'ep': r['ep'],
            'pp': r['pp'], 'ytd': r['ytd'],
        })

    # Sort each stock's history by quarter
    for s in by_stock.values():
        s['hist'].sort(key=lambda x: sort_quarter_key(x['q']))

    # industry aggregation per quarter
    by_ind = defaultdict(lambda: defaultdict(lambda: {'pv': 0.0, 'fc_sum': 0, 'cnt': 0}))
    for r in records:
        if r['pv'] and r['ind']:
            slot = by_ind[r['q']][r['ind']]
            slot['pv'] += r['pv']
            slot['fc_sum'] += (r['fc'] or 0)
            slot['cnt'] += 1

    # Convert to plain dicts
    by_ind_plain = {}
    for q, inds in by_ind.items():
        by_ind_plain[q] = {ind: v for ind, v in inds.items()}

    # ── 加仓幅度三分位（按季度内 fcc>0 的股票三等分）────────────────────
    # tier: 'L'低加仓 / 'M'中加仓 / 'H'高加仓 / 'R'减仓 / 'N'持平
    for q in quarters:
        recs = by_quarter[q]
        fcc_pos = sorted([r['fcc'] for r in recs
                          if r.get('fcc') is not None and r['fcc'] > 0])
        if len(fcc_pos) >= 3:
            t33 = fcc_pos[len(fcc_pos) // 3]
            t66 = fcc_pos[2 * len(fcc_pos) // 3]
        else:
            t33 = t66 = float('inf')
        for r in recs:
            fcc = r.get('fcc')
            if fcc is None:       r['tier'] = 'N'
            elif fcc < 0:         r['tier'] = 'R'
            elif fcc == 0:        r['tier'] = 'N'
            elif fcc <= t33:      r['tier'] = 'L'
            elif fcc <= t66:      r['tier'] = 'M'
            else:                 r['tier'] = 'H'

    # ── 行业平均 fw（占基金持仓比重），用于判断行业低配 ─────────────────
    ind_avg_fw = {}   # {q: {ind: avg_fw}}
    for q in quarters:
        bucket = {}
        for r in by_quarter[q]:
            ind = r.get('ind', '其他')
            fw  = r.get('fw')
            if fw is not None:
                bucket.setdefault(ind, []).append(fw)
        ind_avg_fw[q] = {ind: round(sum(v) / len(v), 6)
                         for ind, v in bucket.items()}
        # 写回每条记录
        for r in by_quarter[q]:
            r['ifw'] = ind_avg_fw[q].get(r.get('ind', '其他'))

    return {
        'quarters': quarters,
        'by_quarter': {q: by_quarter[q] for q in quarters},
        'by_stock': by_stock,
        'by_ind': by_ind_plain,
    }


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>基金重仓看板</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
<style>
:root {
  --bg:         #f0ece4;
  --bg-card:    #ffffff;
  --bg-th:      #f7f5f1;
  --bg-alt:     #faf8f5;
  --bg-hdr:     #14213d;
  --bg-nav:     #ffffff;
  --text-1:     #14213d;
  --text-2:     #566070;
  --text-3:     #96a0ae;
  --accent:     #1e50a2;
  --accent-sub: #eef3fc;
  --pos:        #c0392b;
  --pos-bg:     #fdf2f1;
  --neg:        #166534;
  --neg-bg:     #f0f8f3;
  --bd:         #ddd8d0;
  --bd-th:      #ccc8bf;
  --sh-sm:      0 1px 3px rgba(20,33,61,.07),0 1px 2px rgba(20,33,61,.04);
  --sh:         0 2px 8px rgba(20,33,61,.09),0 1px 3px rgba(20,33,61,.05);
  --r:          8px;
  --mono:       'JetBrains Mono','Courier New',monospace;
  --ui:         'Sora','PingFang SC','Microsoft YaHei',sans-serif;
}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--text-1);font-family:var(--ui);font-size:13px;line-height:1.5}
::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:var(--bd);border-radius:4px}
::-webkit-scrollbar-thumb:hover{background:#bbb5ab}

/* ── Header ── */
.hdr{background:var(--bg-hdr);padding:0 24px;height:52px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:50;box-shadow:0 1px 0 rgba(255,255,255,.08)}
.hdr-brand{display:flex;align-items:center;gap:10px}
.hdr-icon{width:32px;height:32px;background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.18);border-radius:7px;display:flex;align-items:center;justify-content:center;flex-shrink:0}
.hdr h1{font-size:14px;font-weight:700;color:#fff;letter-spacing:.2px;line-height:1.2}
.hdr-sub{font-size:9px;color:rgba(255,255,255,.38);letter-spacing:1.2px;font-weight:400;text-transform:uppercase}
.hdr .meta{color:rgba(255,255,255,.35);font-size:11px;font-family:var(--mono);letter-spacing:.2px}

/* ── Nav ── */
.tabs{background:var(--bg-nav);padding:0 24px;display:flex;border-bottom:1px solid var(--bd);position:sticky;top:52px;z-index:40;box-shadow:var(--sh-sm)}
.tab-btn{padding:14px 18px;cursor:pointer;border:none;background:none;color:var(--text-3);font-size:12px;font-family:var(--ui);font-weight:500;border-bottom:2px solid transparent;transition:all .18s;white-space:nowrap;margin-bottom:-1px;letter-spacing:.1px}
.tab-btn:hover{color:var(--text-2)}
.tab-btn.active{color:var(--accent);border-bottom-color:var(--accent);font-weight:700}

/* ── Content ── */
.tab-pane{display:none;padding:20px 24px;min-height:calc(100vh - 101px)}
.tab-pane.active{display:block}

/* ── Controls ── */
.ctrl{display:flex;gap:8px;align-items:center;margin-bottom:16px;flex-wrap:wrap}
.ctrl-lbl{color:var(--text-3);font-size:10px;font-weight:600;white-space:nowrap;letter-spacing:.6px;text-transform:uppercase}
select,input[type=text]{background:var(--bg-card);border:1px solid var(--bd);color:var(--text-1);padding:5px 10px;border-radius:6px;font-size:12px;font-family:var(--ui);outline:none;height:30px;transition:border .15s,box-shadow .15s}
select:focus,input:focus{border-color:var(--accent);box-shadow:0 0 0 3px rgba(30,80,162,.1)}

/* ── Cards ── */
.cards{display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap}
.card{background:var(--bg-card);border:1px solid var(--bd);border-radius:var(--r);padding:12px 16px;min-width:140px;box-shadow:var(--sh-sm)}
.card-lbl{color:var(--text-3);font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:.7px}
.card-val{font-size:22px;font-weight:700;color:var(--text-1);margin-top:4px;font-family:var(--mono);letter-spacing:-.5px;line-height:1}
.card-sub{font-size:11px;color:var(--text-3);margin-top:3px}

/* ── Table ── */
.tbl-wrap{overflow-x:auto;border-radius:var(--r);border:1px solid var(--bd);box-shadow:var(--sh-sm);background:var(--bg-card)}
table{width:100%;border-collapse:collapse}
thead{position:sticky;top:0;z-index:10}
th{background:var(--bg-th);color:var(--text-3);font-weight:700;font-size:10px;text-transform:uppercase;letter-spacing:.5px;padding:10px 12px;text-align:right;white-space:nowrap;cursor:pointer;user-select:none;border-bottom:1px solid var(--bd-th)}
th:first-child,th.l{text-align:left}
th:hover{color:var(--accent)}
th.asc::after{content:' ↑';color:var(--accent)}
th.desc::after{content:' ↓';color:var(--accent)}
td{padding:8px 12px;border-top:1px solid #ede9e3;text-align:right;white-space:nowrap;font-size:12px;font-family:var(--mono);color:var(--text-1)}
td.l{text-align:left;font-family:var(--ui);font-size:13px}
tr:nth-child(even) td{background:var(--bg-alt)}
tr:hover td{background:var(--accent-sub) !important}
.pos{color:var(--pos);font-weight:600}
.neg{color:var(--neg);font-weight:600}
.neu{color:var(--text-3)}
.code-cell{color:var(--accent);font-family:var(--mono);font-size:11px;font-weight:500}
.rank{color:var(--text-3);font-size:11px;font-family:var(--mono)}
.ind-tag{display:inline-block;background:var(--accent-sub);color:var(--accent);padding:2px 7px;border-radius:4px;font-size:11px;font-family:var(--ui);font-weight:600}
/* 加仓三分位标签 */
.tier{display:inline-block;padding:1px 5px;border-radius:3px;font-size:10px;font-weight:600;font-family:var(--ui);margin-left:4px;vertical-align:middle}
.tier-L{background:#eef3fc;color:#1e50a2}
.tier-M{background:#f5f5f5;color:#96a0ae}
.tier-H{background:#fff3e0;color:#c07000}
.tier-R{background:#fdf2f1;color:#c0392b}
/* 行业低配标签 */
.ifw-low{display:inline-block;padding:1px 5px;border-radius:3px;font-size:10px;font-weight:600;font-family:var(--ui);background:#f0f8f3;color:#166534;margin-left:3px;vertical-align:middle}
.ifw-ref{font-size:9.5px;color:var(--text-3);font-family:var(--mono);display:block;margin-top:1px}
/* 减仓行提示 */
tr.jianc-row td{background:#fff8f7!important}
tr.jianc-row:hover td{background:#fdf2f1!important}

/* ── Charts ── */
.chart-box{background:var(--bg-card);border:1px solid var(--bd);border-radius:var(--r);padding:14px 16px;box-shadow:var(--sh-sm)}
.chart-title{font-size:9px;color:var(--text-3);margin-bottom:10px;font-weight:700;text-transform:uppercase;letter-spacing:.7px}

/* ── Grids ── */
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.grid3{display:grid;grid-template-columns:3fr 2fr;gap:14px}

/* ── Tab3 Search ── */
.search-wrap{position:relative;display:inline-block}
.search-wrap input{width:280px}
.search-dd{position:absolute;top:calc(100% + 4px);left:0;min-width:100%;background:var(--bg-card);border:1px solid var(--bd);border-radius:8px;max-height:240px;overflow-y:auto;z-index:200;display:none;box-shadow:var(--sh)}
.search-dd.show{display:block}
.s-item{padding:8px 14px;cursor:pointer;display:flex;gap:10px;align-items:center;transition:background .1s}
.s-item:hover{background:var(--bg-alt)}
.s-code{color:var(--accent);font-family:var(--mono);font-size:11px;font-weight:500;min-width:88px}
.s-name{color:var(--text-1);font-size:13px}
.s-ind{color:var(--text-3);font-size:11px}
</style>
</head>
<body>

<div class="hdr">
  <div class="hdr-brand">
    <div class="hdr-icon">
      <svg width="16" height="16" viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg">
        <rect x="1" y="9" width="3" height="6" rx="1" fill="white" fill-opacity=".85"/>
        <rect x="6" y="5" width="3" height="10" rx="1" fill="white"/>
        <rect x="11" y="2" width="3" height="13" rx="1" fill="white" fill-opacity=".85"/>
      </svg>
    </div>
    <div>
      <h1>基金重仓看板</h1>
      <div class="hdr-sub">Fund Holdings Monitor</div>
    </div>
  </div>
  <span class="meta" id="meta-info"></span>
</div>

<div class="tabs">
  <button class="tab-btn active" onclick="switchTab('t1')">机构集中度</button>
  <button class="tab-btn" onclick="switchTab('t2')">季度增减持</button>
  <button class="tab-btn" onclick="switchTab('t3')">个股历史趋势</button>
  <button class="tab-btn" onclick="switchTab('t4')">行业分布</button>
</div>

<!-- TAB 1: 机构集中度 -->
<div id="t1" class="tab-pane active">
  <div class="ctrl">
    <span class="ctrl-lbl">季度</span>
    <select id="t1-q"></select>
    <span class="ctrl-lbl">排序</span>
    <select id="t1-sort">
      <option value="fw">占基金持仓</option>
      <option value="pv">持仓市值</option>
      <option value="fp">占流通股比</option>
    </select>
    <span class="ctrl-lbl">行业</span>
    <select id="t1-ind" style="min-width:90px"></select>
    <span class="ctrl-lbl">配置</span>
    <select id="t1-ifw">
      <option value="">全部</option>
      <option value="low">行业低配</option>
      <option value="high">行业高配</option>
    </select>
    <span class="ctrl-lbl">显示</span>
    <select id="t1-top">
      <option value="50">Top 50</option>
      <option value="100">Top 100</option>
      <option value="200">Top 200</option>
      <option value="9999">全部</option>
    </select>
  </div>
  <div id="t1-cards" class="cards"></div>
  <div class="tbl-wrap">
    <table id="t1-tbl">
      <thead>
        <tr>
          <th class="l" style="width:36px">#</th>
          <th class="l" onclick="t1Sort('code')">代码</th>
          <th class="l" onclick="t1Sort('name')">名称</th>
          <th class="l" onclick="t1Sort('ind')">行业</th>
          <th onclick="t1Sort('pv')">持仓市值(亿)</th>
          <th onclick="t1Sort('vc')">市值变动(亿)</th>
          <th onclick="t1Sort('fp')">占流通股比</th>
          <th onclick="t1Sort('fw')">占基金持仓</th>
          <th onclick="t1Sort('fwd')" title="较上季度变动（百分点）">环比变动(pp)</th>
          <th onclick="t1Sort('fc')">持有基金数</th>
          <th onclick="t1Sort('fcc')">本季增减 / 信号</th>
        </tr>
      </thead>
      <tbody id="t1-body"></tbody>
    </table>
  </div>
</div>

<!-- TAB 2: 增减持 -->
<div id="t2" class="tab-pane">
  <div class="ctrl">
    <span class="ctrl-lbl">季度</span>
    <select id="t2-q"></select>
    <span class="ctrl-lbl">指标</span>
    <select id="t2-metric">
      <option value="fwd">占基金持仓环比</option>
      <option value="fw">占基金持仓</option>
      <option value="fcc">基金增减数量</option>
    </select>
    <span class="ctrl-lbl">数量</span>
    <select id="t2-n">
      <option value="20">Top 20</option>
      <option value="30">Top 30</option>
    </select>
  </div>
  <div class="grid2" style="margin-bottom:14px">
    <div class="chart-box">
      <div class="chart-title">加仓 Top</div>
      <div id="chart-add" style="height:400px"></div>
    </div>
    <div class="chart-box">
      <div class="chart-title">减仓 Top</div>
      <div id="chart-red" style="height:400px"></div>
    </div>
  </div>
</div>

<!-- TAB 3: 个股历史 -->
<div id="t3" class="tab-pane">
  <div class="ctrl">
    <span class="ctrl-lbl">搜索股票</span>
    <div class="search-wrap">
      <input type="text" id="t3-search" placeholder="代码 / 名称" autocomplete="off">
      <div id="t3-dd" class="search-dd"></div>
    </div>
  </div>
  <div id="t3-detail" style="display:none">
    <div class="cards" id="t3-cards"></div>
    <div class="grid2" style="margin-bottom:14px">
      <div class="chart-box">
        <div class="chart-title">持仓市值(亿元)</div>
        <div id="chart-pv" style="height:260px"></div>
      </div>
      <div class="chart-box">
        <div class="chart-title">占基金持仓比重(%)</div>
        <div id="chart-fw" style="height:260px"></div>
      </div>
    </div>
    <div class="grid2" style="margin-bottom:14px">
      <div class="chart-box">
        <div class="chart-title">占流通股比(%)</div>
        <div id="chart-fp" style="height:220px"></div>
      </div>
      <div class="chart-box">
        <div class="chart-title">持有基金数</div>
        <div id="chart-fc" style="height:220px"></div>
      </div>
    </div>
    <div class="tbl-wrap">
      <table><thead><tr>
        <th class="l">季度</th>
        <th>季末股价</th>
        <th>持仓数量(万股)</th>
        <th>持仓变动(万股)</th>
        <th>持仓市值(亿)</th>
        <th>持有基金数</th>
        <th>本季增减</th>
        <th>占流通股比</th>
        <th>年初至今</th>
      </tr></thead>
      <tbody id="t3-body"></tbody>
      </table>
    </div>
  </div>
  <div id="t3-empty" style="color:#555;padding:40px 0;text-align:center">搜索股票代码或名称查看历史趋势</div>
</div>

<!-- TAB 4: 行业分布 -->
<div id="t4" class="tab-pane">
  <div class="ctrl">
    <span class="ctrl-lbl">季度</span>
    <select id="t4-q"></select>
    <span class="ctrl-lbl">指标</span>
    <select id="t4-metric">
      <option value="pv">持仓市值(亿)</option>
      <option value="cnt">持仓股票数</option>
    </select>
  </div>
  <div class="grid3" style="margin-bottom:14px">
    <div class="chart-box">
      <div class="chart-title">行业持仓分布</div>
      <div id="chart-ind-bar" style="height:400px"></div>
    </div>
    <div class="chart-box">
      <div class="chart-title">持仓比例</div>
      <div id="chart-ind-pie" style="height:300px"></div>
      <div id="t4-pie-legend" style="margin-top:10px;padding-top:10px;border-top:1px solid #ede9e3"></div>
    </div>
  </div>
  <div class="chart-box" style="margin-bottom:14px">
    <div class="chart-title">行业配置跨季度变化（持仓市值亿元）</div>
    <div id="chart-ind-trend" style="height:320px"></div>
  </div>
  <div class="tbl-wrap">
    <table id="t4-tbl">
      <thead><tr>
        <th class="l">行业</th>
        <th>持仓市值(亿)</th>
        <th>持仓股票数</th>
        <th>平均持有基金数</th>
      </tr></thead>
      <tbody id="t4-body"></tbody>
    </table>
  </div>
</div>

<script>
const DATA = __DATA__;

// ── helpers ──────────────────────────────────────────────────────────
const qs = DATA.quarters;
const latestQ = qs[qs.length - 1];

function fmt(v, digits=2) {
  if (v == null || isNaN(v)) return '—';
  return v.toFixed(digits);
}
function fmtPct(v) {
  if (v == null || isNaN(v)) return '—';
  return (v * 100).toFixed(2) + '%';
}
function fmtPctRaw(v) {
  // v already in percent form
  if (v == null || isNaN(v)) return '—';
  return v.toFixed(2) + '%';
}
function clsPct(v) {
  if (v == null) return 'neu';
  return v > 0 ? 'pos' : v < 0 ? 'neg' : 'neu';
}
function clsInt(v) {
  if (v == null) return 'neu';
  return v > 0 ? 'pos' : v < 0 ? 'neg' : 'neu';
}
function sign(v) {
  if (v == null || isNaN(v)) return '';
  return v > 0 ? '+' : '';
}

// Light-theme axis/grid defaults for ECharts
const AX = {
  axisLine: { lineStyle: { color: '#ccc8bf' } },
  axisLabel: { color: '#96a0ae', fontSize: 10, fontFamily: "'JetBrains Mono',monospace" },
  splitLine: { lineStyle: { color: '#e8e3db', type: 'dashed' } },
};
const TT = { backgroundColor:'#fff', borderColor:'#ddd8d0', borderWidth:1, textStyle:{color:'#14213d',fontSize:11} };
// Categorical palette for multi-series charts
const CAT = ['#1e50a2','#c0392b','#166534','#d97706','#7c3aed','#0891b2','#be185d','#0d9488','#ea580c','#4f46e5'];

function makeChart(id) {
  const el = document.getElementById(id);
  if (!el) return null;
  let c = echarts.getInstanceByDom(el);
  if (c) c.dispose();
  return echarts.init(el, null, { renderer: 'canvas' });
}

// ── tab switching ─────────────────────────────────────────────────────
let t3DefaultRendered = false;
function switchTab(id) {
  document.querySelectorAll('.tab-pane').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  const idx = ['t1','t2','t3','t4'].indexOf(id);
  document.querySelectorAll('.tab-btn')[idx].classList.add('active');
  if (id === 't2') renderT2();
  if (id === 't4') renderT4();
  if (id === 't3' && !t3DefaultRendered && DATA.by_stock['9988.HK']) {
    t3DefaultRendered = true;
    selectStock('9988.HK');
  }
}

// ── populate selects ─────────────────────────────────────────────────
function populateQ(id, onChange) {
  const sel = document.getElementById(id);
  qs.forEach(q => {
    const opt = document.createElement('option');
    opt.value = q; opt.text = q;
    sel.appendChild(opt);
  });
  sel.value = latestQ;
  if (onChange) sel.addEventListener('change', onChange);
}

// ── TAB 1: 机构集中度 ────────────────────────────────────────────────
let t1SortKey = 'fw', t1SortAsc = false;

function t1Sort(key) {
  if (t1SortKey === key) t1SortAsc = !t1SortAsc;
  else { t1SortKey = key; t1SortAsc = false; }
  renderT1();
}

function populateT1Industries(q) {
  const sel = document.getElementById('t1-ind');
  const cur = sel.value;
  const inds = [...new Set((DATA.by_quarter[q]||[]).map(r=>r.ind).filter(Boolean))].sort();
  sel.innerHTML = '<option value="">全部行业</option>' +
    inds.map(i => `<option value="${i}"${i===cur?'selected':''}>${i}</option>`).join('');
}

function renderT1() {
  const q = document.getElementById('t1-q').value;
  const sortBy = document.getElementById('t1-sort').value;
  const topN = parseInt(document.getElementById('t1-top').value);
  const indFilter = document.getElementById('t1-ind').value;
  const ifwFilter = document.getElementById('t1-ifw').value;
  let rows = (DATA.by_quarter[q] || []).slice();
  if (indFilter) rows = rows.filter(r => r.ind === indFilter);
  if (ifwFilter === 'low')  rows = rows.filter(r => r.fw != null && r.ifw != null && r.fw < r.ifw);
  if (ifwFilter === 'high') rows = rows.filter(r => r.fw != null && r.ifw != null && r.fw >= r.ifw);

  // 计算 fw 环比变动（百分点）
  const qIdx = qs.indexOf(q);
  const prevQ = qIdx > 0 ? qs[qIdx - 1] : null;
  const prevFwMap = {};
  if (prevQ) (DATA.by_quarter[prevQ] || []).forEach(r => { prevFwMap[r.code] = r.fw; });
  rows = rows.map(r => ({
    ...r,
    fwd: (prevQ && r.fw != null && prevFwMap[r.code] != null)
      ? +((r.fw - prevFwMap[r.code]) * 100).toFixed(4)
      : null
  }));

  // Use UI sort override if user clicked column header, else use dropdown
  const sk = t1SortKey || sortBy;
  rows.sort((a, b) => {
    const av = a[sk], bv = b[sk];
    if (av == null && bv == null) return 0;
    if (av == null) return 1;
    if (bv == null) return -1;
    return t1SortAsc ? av - bv : bv - av;
  });

  // Summary cards
  const total = rows.length;
  const totalPv = rows.reduce((s,r) => s + (r.pv||0), 0);
  const topStock = rows[0];
  const topFw = topStock?.fw != null ? (topStock.fw*100).toFixed(3)+'%' : '—';
  document.getElementById('t1-cards').innerHTML = `
    <div class="card"><div class="ctrl-lbl">季度</div><div class="card-val" style="font-size:20px">${q}</div></div>
    <div class="card"><div class="ctrl-lbl">持仓股票数</div><div class="card-val">${total}</div></div>
    <div class="card"><div class="ctrl-lbl">持仓市值合计</div><div class="card-val">${totalPv.toFixed(0)}亿</div></div>
    <div class="card"><div class="ctrl-lbl">占基金持仓最高</div><div class="card-val" style="font-size:16px">${topFw}</div><div class="card-sub">${topStock?.name||''}</div></div>
  `;

  // Update header sort indicators
  document.querySelectorAll('#t1-tbl th').forEach(th => {
    th.classList.remove('asc','desc');
    const fn = th.getAttribute('onclick');
    if (fn && fn.includes(`'${t1SortKey}'`)) {
      th.classList.add(t1SortAsc ? 'asc' : 'desc');
    }
  });

  const top = rows.slice(0, topN);
  const tbody = document.getElementById('t1-body');

  // 加仓三分位标签配置
  const TIER_CFG = {
    L: ['tier-L', '低加仓'],
    M: ['tier-M', '中加仓'],
    H: ['tier-H', '高加仓⚠'],
    R: ['tier-R', '减仓⚠'],
    N: ['tier-M', '持平'],
  };

  tbody.innerHTML = top.map((r, i) => {
    const fccCls  = clsInt(r.fcc);
    const vcCls   = clsPct(r.vc);
    const isJianc = r.fcc != null && r.fcc < 0;

    // 占基金持仓 + 行业均值
    let fwCell = '—';
    if (r.fw != null) {
      const fwPct  = (r.fw * 100).toFixed(3) + '%';
      const ifwPct = r.ifw != null ? (r.ifw * 100).toFixed(3) + '%' : null;
      const isLow  = r.ifw != null && r.fw < r.ifw;
      const lowBadge = isLow ? '<span class="ifw-low" title="低于行业平均配置">低配</span>' : '';
      const ifwRef   = ifwPct ? `<span class="ifw-ref">行业均 ${ifwPct}</span>` : '';
      fwCell = `${fwPct}${lowBadge}${ifwRef}`;
    }

    // 本季增减 + 三分位标签
    let fccCell = '—';
    if (r.fcc != null) {
      const tier = r.tier || 'N';
      const [tCls, tLbl] = TIER_CFG[tier] || TIER_CFG['N'];
      fccCell = `${sign(r.fcc)}${r.fcc}<span class="tier ${tCls}">${tLbl}</span>`;
    }

    return `<tr class="${isJianc ? 'jianc-row' : ''}">
      <td class="l rank">${i+1}</td>
      <td class="l code-cell">${r.code}</td>
      <td class="l">${r.name}</td>
      <td class="l"><span class="ind-tag">${r.ind}</span></td>
      <td>${fmt(r.pv)}</td>
      <td class="${vcCls}">${r.vc != null ? sign(r.vc)+fmt(r.vc) : '—'}</td>
      <td>${r.fp != null ? (r.fp*100).toFixed(2)+'%' : '—'}</td>
      <td class="l">${fwCell}</td>
      <td class="${clsPct(r.fwd)}">${r.fwd != null ? sign(r.fwd)+r.fwd.toFixed(3)+'pp' : '—'}</td>
      <td>${r.fc != null ? r.fc : '—'}</td>
      <td class="${fccCls} l">${fccCell}</td>
    </tr>`;
  }).join('');
}

// ── TAB 2: 增减持 ────────────────────────────────────────────────────
let chartAdd, chartRed;

function renderT2() {
  const q = document.getElementById('t2-q').value;
  const metric = document.getElementById('t2-metric').value;
  const n = parseInt(document.getElementById('t2-n').value);
  // 计算 fwd（占基金持仓环比变动，百分点），需在 filter 之前完成
  const qIdx = qs.indexOf(q);
  const prevQ2 = qIdx > 0 ? qs[qIdx - 1] : null;
  const prevFwMap2 = {};
  if (prevQ2) (DATA.by_quarter[prevQ2] || []).forEach(r => { prevFwMap2[r.code] = r.fw; });
  let rows = (DATA.by_quarter[q] || []).map(r => ({
    ...r,
    fwd: (prevQ2 && r.fw != null && prevFwMap2[r.code] != null)
      ? +((r.fw - prevFwMap2[r.code]) * 100).toFixed(4)
      : null
  })).filter(r => r[metric] != null);

  // fw 是绝对值指标：显示重仓/轻仓；其他是变化量指标：显示增/减
  const isFw = metric === 'fw';
  const labelMap = { fw:'占基金持仓', fcc:'基金增减数量', fwd:'占基金持仓环比(pp)' };
  const label = labelMap[metric] || metric;

  let adds, reds, titleAdd, titleRed;
  if (isFw) {
    adds = rows.slice().sort((a,b) => b.fw - a.fw).slice(0, n);
    reds = rows.slice().sort((a,b) => a.fw - b.fw).filter(r=>r.fw>0).slice(0, n);
    titleAdd = '重仓 Top'; titleRed = '轻仓 Top';
  } else {
    adds = rows.filter(r => r[metric] > 0).sort((a,b) => b[metric]-a[metric]).slice(0,n);
    reds = rows.filter(r => r[metric] < 0).sort((a,b) => a[metric]-b[metric]).slice(0,n);
    titleAdd = metric === 'fwd' ? '提升 Top' : '加仓 Top';
    titleRed  = metric === 'fwd' ? '下降 Top' : '减仓 Top';
  }
  // Update section titles
  document.querySelector('#chart-add').closest('.chart-box').querySelector('.chart-title').textContent = titleAdd;
  document.querySelector('#chart-red').closest('.chart-box').querySelector('.chart-title').textContent = titleRed;

  function barOpt(data, color, reversed) {
    const isFwd = metric === 'fwd';
    return {
      backgroundColor: 'transparent',
      grid: { left: 120, right: 60, top: 10, bottom: 30 },
      xAxis: { type:'value', ...AX, axisLine:{show:false} },
      yAxis: {
        type:'category', data: data.map(r => r.name), inverse: !reversed,
        axisLine:{show:false}, axisTick:{show:false},
        axisLabel: { color:'#566070', fontSize:11, width:112, overflow:'truncate', fontFamily:"'Sora','PingFang SC',sans-serif" }
      },
      tooltip: { ...TT, trigger:'axis',
        formatter: p => {
          const r = data[p[0].dataIndex];
          const val = isFw ? (r.fw*100).toFixed(3)+'%'
                    : isFwd ? r.fwd.toFixed(3)+'pp'
                    : r[metric];
          return `<b>${r.code} ${r.name}</b><br>${label}: ${val}<br>持有基金数: ${r.fc}`;
        }
      },
      series: [{
        type:'bar',
        data: data.map(r => isFw ? +(r.fw*100).toFixed(4) : r[metric]),
        barMaxWidth: 28,
        itemStyle: { color, borderRadius:[0,4,4,0] },
        label: { show:true, position:'right', color:'#96a0ae', fontSize:10, fontFamily:"'JetBrains Mono',monospace",
          formatter: p => isFw ? p.value.toFixed(3)+'%'
                       : isFwd ? Math.abs(p.value).toFixed(3)+'pp'
                       : Math.abs(p.value) }
      }]
    };
  }

  chartAdd = makeChart('chart-add');
  if (chartAdd) chartAdd.setOption(barOpt(adds, '#c0392b', false));

  chartRed = makeChart('chart-red');
  if (chartRed) chartRed.setOption(barOpt(reds, '#166534', true));
}

// ── TAB 3: 个股历史 ──────────────────────────────────────────────────
let t3Charts = [];

function initT3Search() {
  const inp = document.getElementById('t3-search');
  const dd = document.getElementById('t3-dd');
  const stocks = Object.entries(DATA.by_stock);

  inp.addEventListener('input', () => {
    const q = inp.value.trim().toLowerCase();
    if (!q) { dd.classList.remove('show'); return; }
    const hits = stocks.filter(([code, s]) =>
      code.toLowerCase().includes(q) || s.name.includes(q)
    ).slice(0,12);
    if (!hits.length) { dd.classList.remove('show'); return; }
    dd.innerHTML = hits.map(([code, s]) =>
      `<div class="s-item" onclick="selectStock('${code}')">
        <span class="s-code">${code}</span>
        <span class="s-name">${s.name}</span>
        <span class="s-ind">${s.ind}</span>
      </div>`
    ).join('');
    dd.classList.add('show');
  });

  document.addEventListener('click', e => {
    if (!e.target.closest('.search-wrap')) dd.classList.remove('show');
  });
}

function selectStock(code) {
  const s = DATA.by_stock[code];
  if (!s) return;
  document.getElementById('t3-search').value = `${code}  ${s.name}`;
  document.getElementById('t3-dd').classList.remove('show');
  renderStockDetail(code, s);
}

function renderStockDetail(code, s) {
  const hist = s.hist;
  const qs = hist.map(h => h.q);
  const detail = document.getElementById('t3-detail');
  const empty = document.getElementById('t3-empty');
  detail.style.display = 'block';
  empty.style.display = 'none';

  // Summary cards
  const latest = hist[hist.length - 1];
  const prev = hist.length > 1 ? hist[hist.length - 2] : null;
  const fcChg = prev && latest.fc != null && prev.fc != null ? latest.fc - prev.fc : null;
  const latestFcc = latest.fcc;
  const jianCSignal = latestFcc != null && latestFcc < 0;
  const jianCHtml = jianCSignal
    ? `<div style="background:#fdf2f1;border:1px solid #f5c6c2;border-radius:6px;
        padding:8px 14px;margin-bottom:12px;font-size:12px;color:#c0392b;font-weight:600">
        ⚠ 减仓信号（${latest.q}）：持有基金数净减少 ${latestFcc} 家，机构正在撤出，建议关注。
       </div>` : '';
  document.getElementById('t3-cards').innerHTML = jianCHtml + `
    <div class="card"><div class="ctrl-lbl">代码</div><div class="card-val" style="font-size:16px">${code}</div><div class="card-sub">${s.name} · ${s.ind}</div></div>
    <div class="card"><div class="ctrl-lbl">最新季度</div><div class="card-val" style="font-size:18px">${latest.q}</div></div>
    <div class="card"><div class="ctrl-lbl">持有基金数</div><div class="card-val">${latest.fc??'—'}</div><div class="card-sub ${clsInt(fcChg)}">${fcChg!=null?(sign(fcChg)+fcChg+'家'):'—'}</div></div>
    <div class="card"><div class="ctrl-lbl">持仓市值</div><div class="card-val">${latest.pv!=null?latest.pv.toFixed(1)+'亿':'—'}</div></div>
    <div class="card"><div class="ctrl-lbl">占流通股比</div><div class="card-val">${latest.fp!=null?(latest.fp*100).toFixed(2)+'%':'—'}</div></div>
  `;

  // Dispose old charts
  t3Charts.forEach(c => c && c.dispose());
  t3Charts = [];

  function lineOpt(series, yName, color) {
    return {
      backgroundColor: 'transparent',
      grid: { left:54, right:16, top:20, bottom:34 },
      xAxis: { type:'category', data:qs, ...AX, axisTick:{show:false} },
      yAxis: { type:'value', name:yName, nameTextStyle:{color:'#96a0ae',fontSize:9,padding:[0,0,0,0]}, ...AX, axisLine:{show:false} },
      tooltip: { ...TT, trigger:'axis' },
      series: [{ type:'line', data:series, smooth:.4,
        itemStyle:{color, borderColor:'#fff', borderWidth:1.5},
        lineStyle:{color,width:2.5},
        areaStyle:{color:{type:'linear',x:0,y:0,x2:0,y2:1,
          colorStops:[{offset:0,color:color+'28'},{offset:1,color:color+'04'}]}},
        symbolSize:6 }]
    };
  }

  function barOpt2(series) {
    return {
      backgroundColor: 'transparent',
      grid: { left:54, right:16, top:20, bottom:34 },
      xAxis: { type:'category', data:qs, ...AX, axisTick:{show:false} },
      yAxis: { type:'value', ...AX, axisLine:{show:false} },
      tooltip: { ...TT, trigger:'axis' },
      series: [{ type:'bar', data:series, barMaxWidth:32,
        itemStyle: { color: p => p.value >= 0 ? '#c0392b' : '#166534', borderRadius:[3,3,0,0] } }]
    };
  }

  const cpv = makeChart('chart-pv');
  cpv && cpv.setOption(lineOpt(hist.map(h=>h.pv), '亿元', '#d97706'));
  t3Charts.push(cpv);

  const cfw = makeChart('chart-fw');
  cfw && cfw.setOption(lineOpt(hist.map(h=>h.fw!=null?+(h.fw*100).toFixed(4):null), '%', '#1e50a2'));
  t3Charts.push(cfw);

  const cfp = makeChart('chart-fp');
  cfp && cfp.setOption(lineOpt(hist.map(h=>h.fp!=null?+(h.fp*100).toFixed(3):null), '%', '#7c3aed'));
  t3Charts.push(cfp);

  const cfc = makeChart('chart-fc');
  cfc && cfc.setOption(lineOpt(hist.map(h=>h.fc), '', '#0891b2'));
  t3Charts.push(cfc);

  // Table
  document.getElementById('t3-body').innerHTML = hist.slice().reverse().map(h => {
    const isJianc = h.fcc != null && h.fcc < 0;
    const fccLabel = h.fcc != null
      ? `${sign(h.fcc)}${h.fcc}${isJianc ? ' <span class="tier tier-R" title="减仓信号">减仓⚠</span>' : ''}`
      : '—';
    return `<tr class="${isJianc ? 'jianc-row' : ''}">
    <td class="l"><b>${h.q}</b></td>
    <td>${fmt(h.ep)}</td>
    <td>${h.sh!=null?fmt(h.sh):'—'}</td>
    <td class="${clsPct(h.sch)}">${h.sch!=null?sign(h.sch)+fmt(h.sch):'—'}</td>
    <td>${h.pv!=null?fmt(h.pv):'—'}</td>
    <td><b>${h.fc!=null?h.fc:'—'}</b></td>
    <td class="${clsInt(h.fcc)} l">${fccLabel}</td>
    <td>${h.fp!=null?(h.fp*100).toFixed(2)+'%':'—'}</td>
    <td class="${clsPct(h.ytd)}">${h.ytd!=null?sign(h.ytd)+fmtPct(h.ytd):'—'}</td>
  </tr>`;
  }).join('');
}

// ── TAB 4: 行业分布 ──────────────────────────────────────────────────
function renderT4() {
  const q = document.getElementById('t4-q').value;
  const metric = document.getElementById('t4-metric').value;
  const qInd = DATA.by_ind[q] || {};

  const sorted = Object.entries(qInd).sort((a,b) => b[1][metric] - a[1][metric]);
  const industries = sorted.map(([ind]) => ind);
  const vals = sorted.map(([,v]) => metric==='pv' ? +v.pv.toFixed(1) : v.cnt);

  // Bar chart — 高度按行业数量动态设置，保证每行有足够空间
  const barH = Math.max(360, industries.length * 28);
  document.getElementById('chart-ind-bar').style.height = barH + 'px';
  const cBar = makeChart('chart-ind-bar');
  cBar && cBar.setOption({
    backgroundColor:'transparent',
    grid:{left:86,right:40,top:10,bottom:36},
    xAxis:{type:'value',...AX,axisLine:{show:false}},
    yAxis:{type:'category',data:industries,inverse:false,axisLine:{show:false},axisTick:{show:false},
      axisLabel:{color:'#566070',fontSize:11,fontFamily:"'Sora','PingFang SC',sans-serif"}},
    tooltip:{...TT,trigger:'axis'},
    series:[{type:'bar',data:vals,barMaxWidth:28,
      itemStyle:{borderRadius:[0,4,4,0],
        color:{type:'linear',x:0,y:0,x2:1,y2:0,
          colorStops:[{offset:0,color:'#1e50a2'},{offset:1,color:'#5b8fde'}]}},
      label:{show:true,position:'right',color:'#96a0ae',fontSize:10,fontFamily:"'JetBrains Mono',monospace"}
    }]
  });

  // Pie chart — 不用内置图例，用下方 HTML 图例
  const pieTotal = vals.reduce((s,v) => s+v, 0);
  const cPie = makeChart('chart-ind-pie');
  cPie && cPie.setOption({
    backgroundColor:'transparent',
    tooltip:{...TT,trigger:'item',formatter:'{b}: {c} ({d}%)'},
    color: CAT,
    series:[{
      type:'pie',radius:['38%','70%'],center:['50%','50%'],
      data: sorted.map(([ind,v]) => ({name:ind, value: metric==='pv'?+v.pv.toFixed(1):v.cnt})),
      label:{show:false},
      emphasis:{itemStyle:{shadowBlur:8,shadowColor:'rgba(20,33,61,.15)'}}
    }]
  });

  // HTML 图例：每行两列，显示色块+行业+占比
  const unit = metric==='pv' ? '亿' : '只';
  document.getElementById('t4-pie-legend').innerHTML =
    `<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px 12px">` +
    sorted.map(([ind,v], i) => {
      const val = metric==='pv' ? +v.pv.toFixed(1) : v.cnt;
      const pct = pieTotal > 0 ? (val/pieTotal*100).toFixed(1) : '0.0';
      const col = CAT[i % CAT.length];
      return `<div style="display:flex;align-items:center;gap:6px;font-size:11px;overflow:hidden">
        <span style="width:9px;height:9px;border-radius:2px;background:${col};flex-shrink:0"></span>
        <span style="color:#566070;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;flex:1">${ind}</span>
        <span style="font-family:'JetBrains Mono',monospace;color:#14213d;white-space:nowrap">${pct}%</span>
      </div>`;
    }).join('') +
    `</div>`;

  // Trend chart (stacked bar, all quarters, top industries by latest Q)
  const topInds = industries.slice(0, 10);
  const cTrend = makeChart('chart-ind-trend');
  cTrend && cTrend.setOption({
    backgroundColor:'transparent',
    color: CAT,
    grid:{left:52,right:130,top:24,bottom:34},
    xAxis:{type:'category',data:qs,...AX,axisTick:{show:false}},
    yAxis:{type:'value',...AX,axisLine:{show:false}},
    tooltip:{...TT,trigger:'axis',axisPointer:{type:'shadow'}},
    legend:{orient:'vertical',right:0,top:'middle',textStyle:{color:'#566070',fontSize:10,fontFamily:"'Sora',sans-serif"},type:'scroll'},
    series: topInds.map((ind, i) => ({
      name: ind, type:'bar', stack:'total',
      data: qs.map(q2 => {
        const v = (DATA.by_ind[q2]||{})[ind];
        return v ? +v.pv.toFixed(1) : 0;
      }),
      emphasis:{focus:'series'}
    }))
  });

  // Table
  document.getElementById('t4-body').innerHTML = sorted.map(([ind, v]) => `<tr>
    <td class="l"><span class="ind-tag">${ind}</span></td>
    <td>${v.pv.toFixed(1)}</td>
    <td>${v.cnt}</td>
    <td>${v.cnt ? (v.fc_sum/v.cnt).toFixed(0) : '—'}</td>
  </tr>`).join('');
}

// ── Init ─────────────────────────────────────────────────────────────
(function init() {
  document.getElementById('meta-info').textContent =
    `数据区间：${qs[0]} ~ ${qs[qs.length-1]}  |  生成时间：__GENTIME__`;

  populateQ('t1-q', () => { populateT1Industries(document.getElementById('t1-q').value); renderT1(); });
  populateT1Industries(latestQ);
  document.getElementById('t1-sort').addEventListener('change', () => { t1SortKey = document.getElementById('t1-sort').value; t1SortAsc = false; renderT1(); });
  document.getElementById('t1-ind').addEventListener('change', renderT1);
  document.getElementById('t1-ifw').addEventListener('change', renderT1);
  document.getElementById('t1-top').addEventListener('change', renderT1);

  populateQ('t2-q', renderT2);
  document.getElementById('t2-metric').addEventListener('change', renderT2);
  document.getElementById('t2-n').addEventListener('change', renderT2);

  populateQ('t4-q', renderT4);
  document.getElementById('t4-metric').addEventListener('change', renderT4);

  initT3Search();
  renderT1();
})();
</script>
</body>
</html>
"""


def build_html(quarters, records):
    dataset = build_dataset(quarters, records)
    data_json = json.dumps(dataset, ensure_ascii=False, separators=(',', ':'))
    gentime = datetime.now().strftime('%Y-%m-%d %H:%M')
    html = HTML_TEMPLATE.replace('__DATA__', data_json).replace('__GENTIME__', gentime)
    return html


def main():
    print("读取 Excel...")
    quarters, records = load_data()
    print(f"  季度: {quarters}")
    print(f"  总记录数: {len(records)}")

    print("生成 HTML...")
    html = build_html(quarters, records)
    OUTPUT.write_text(html, encoding='utf-8')
    size_kb = OUTPUT.stat().st_size / 1024
    print(f"  输出: {OUTPUT}  ({size_kb:.0f} KB)")
    print("完成！用浏览器打开 fund_dashboard.html 查看看板。")


if __name__ == '__main__':
    main()
